# -*- coding: utf-8 -*-
"""
Created on Wed Apr  7 15:40:12 2021

@author: jwehounou
"""

# -*- coding: utf-8 -*-
"""
Created on Wed Mar 17 12:23:32 2021

@author: jwehounou
"""
import os
import time
import psutil

import numpy as np
import pandas as pd
import smartgrids_players as players
import fonctions_auxiliaires as fct_aux
import itertools as it

from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook

global MOD
MOD = 65000 #int(0.10*pow(2, m_players))

###############################################################################
#           definition  des fonctions 
#               * calculant les valeurs des cellules des arrays ou
#               * mettant a jour les cellules des arrays
#
###############################################################################
def update_saving_variables(t, b0_ts_T_BESTBF, b0_t_algo,
                            c0_ts_T_BESTBF, c0_t_algo,
                            BENs_M_T_BESTBF, bens_t_algo,
                            CSTs_M_T_BESTBF, csts_t_algo,
                            pi_sg_plus_T_BESTBF, pi_sg_plus_t_algo,
                            pi_sg_minus_T_BESTBF, pi_sg_minus_t_algo,
                            pi_0_plus_T_BESTBF, pi_0_plus_t_algo,
                            pi_0_minus_T_BESTBF, pi_0_minus_t_algo,
                            df_nash_BESTBF, df_nash_algo):
    b0_ts_T_BESTBF[t] = b0_t_algo
    c0_ts_T_BESTBF[t] = c0_t_algo
    BENs_M_T_BESTBF[:,t] = bens_t_algo
    CSTs_M_T_BESTBF[:,t] = csts_t_algo
    pi_sg_plus_T_BESTBF[t] = pi_sg_plus_t_algo
    pi_sg_minus_T_BESTBF[t] = pi_sg_minus_t_algo
    pi_0_plus_T_BESTBF[t] = pi_0_plus_t_algo
    pi_0_minus_T_BESTBF[t] = pi_0_minus_t_algo
    df_nash_BESTBF = df_nash_algo.copy()
    
    return b0_ts_T_BESTBF, c0_ts_T_BESTBF, \
            BENs_M_T_BESTBF, CSTs_M_T_BESTBF, \
            pi_sg_plus_T_BESTBF, pi_sg_minus_T_BESTBF, \
            pi_0_plus_T_BESTBF, pi_0_minus_T_BESTBF, \
            df_nash_BESTBF
    
    

def compute_prices_variables(arr_pl_M_T_vars_modif_algo, t,
                            b0_ts_T_algo, c0_ts_T_algo,
                            pi_sg_plus_T_algo, pi_sg_minus_T_algo,
                            pi_0_plus_T_algo, pi_0_minus_T_algo):
    # # B_is, C_is of shape (M_PLAYERS, )
    # prod_i_M_T_algo = arr_pl_M_T_vars_modif_algo[
    #                         :,:t_periods, 
    #                         fct_aux.AUTOMATE_INDEX_ATTRS["prod_i"]]
    # cons_i_M_T_algo = arr_pl_M_T_vars_modif_algo[
    #                         :,:t_periods, 
    #                         fct_aux.AUTOMATE_INDEX_ATTRS["cons_i"]]
    # B_is_M_algo = np.sum(b0_ts_T_algo * prod_i_M_T_algo, axis=1)
    # C_is_M_algo = np.sum(c0_ts_T_algo * cons_i_M_T_algo, axis=1)
    
    # # BB_is, CC_is, RU_is of shape (M_PLAYERS, )
    # CONS_is_M_algo = np.sum(cons_i_M_T_algo, axis=1)
    # PROD_is_M_algo = np.sum(prod_i_M_T_algo, axis=1)
    
    # print("{}, t={}, pi_sg_plus_T={}, pi_sg_minus_T={} \n".format(
    #         algo_name, t, pi_sg_plus_T_algo[t], pi_sg_minus_T_algo[t]))
    
    # BB_is_M_algo = pi_sg_plus_T_algo[t] * PROD_is_M_algo #np.sum(PROD_is)
    # for num_pl, bb_i in enumerate(BB_is_M_algo):
    #     if bb_i != 0:
    #         print("player {}, BB_i={}".format(num_pl, bb_i))
    # CC_is_M_algo = pi_sg_minus_T_algo[t] * CONS_is_M_algo #np.sum(CONS_is)
    # RU_is_M_algo = BB_is_M_algo - CC_is_M_algo
    
    # pi_hp_plus_s = np.array([pi_hp_plus] * t_periods, dtype=object)
    # pi_hp_minus_s = np.array([pi_hp_minus] * t_periods, dtype=object)
    
    prod_i_M_T_algo = arr_pl_M_T_vars_modif_algo[
                            :,:t+1, 
                            fct_aux.AUTOMATE_INDEX_ATTRS["prod_i"]] 
    cons_i_M_T_algo = arr_pl_M_T_vars_modif_algo[
                            :,:t+1, 
                            fct_aux.AUTOMATE_INDEX_ATTRS["cons_i"]]
    B_is_M_T_algo = b0_ts_T_algo[:t+1] * prod_i_M_T_algo
    C_is_M_T_algo = c0_ts_T_algo[:t+1] * cons_i_M_T_algo
    BB_is_M_T_algo = pi_sg_plus_T_algo[:t+1] * prod_i_M_T_algo
    CC_is_M_T_algo = pi_sg_minus_T_algo[:t+1] * cons_i_M_T_algo 
    RU_is_M_T_algo = BB_is_M_T_algo - CC_is_M_T_algo

    return BB_is_M_T_algo, CC_is_M_T_algo, RU_is_M_T_algo, \
           B_is_M_T_algo, C_is_M_T_algo
           
           
           
###############################################################################
#                   definition  des fonctions annexes
#
###############################################################################
# __________            find best, bad, mid key --> debut             _________
def select_perf_t_from_key(best=None, current=0, key="best"):
    """
    select the Perf_t respecting the key criter
    key could be best and bad
    """
    best_key_Perf_t = None
    if best is None:
        best_key_Perf_t = current
    elif best >= current and key=="best":
        best_key_Perf_t = best
    elif best < current and key=="best":
        best_key_Perf_t = current
    elif best >= current and key=="bad":
        best_key_Perf_t = current
    elif best < current and key=="bad":
        best_key_Perf_t = best
    
    return best_key_Perf_t

def select_mid_perf_t_from_key(Perf_ts):
    """
    select the median key of Perf_t: mid_key_Perf_t
    """
    return np.argsort(Perf_ts)[len(Perf_ts)//2]
    
# __________            find best, bad, mid key --> fin               _________

# __________            find possibles modes --> debut               _________
def possibles_modes_players_automate(arr_pl_M_t_k_vars, m_players, t=0):
    """
    generate the list of possible modes by the states of players

    Parameters
    ----------
    arr_pl_M_t_k_vars : TYPE, shape (m_players,)
        DESCRIPTION. The default is None.
        it means that t and k are the fixed values 
    t : TYPE, optional
        DESCRIPTION. The default is 0.
    k : TYPE, optional
        DESCRIPTION. The default is 0.

    Returns
    -------
    None.possibles_modes_players_automate

    """

    possibles_modes = list()
    
    for num_pl_i in range(0, m_players):
        state_i = arr_pl_M_t_k_vars[num_pl_i, 
                                    fct_aux.AUTOMATE_INDEX_ATTRS["state_i"]] 
        
        # get mode_i
        if state_i == fct_aux.STATES[0]:                                        # state1 or Deficit
            possibles_modes.append(fct_aux.STATE1_STRATS)
        elif state_i == fct_aux.STATES[1]:                                      # state2 or Self
            possibles_modes.append(fct_aux.STATE2_STRATS)
        elif state_i == fct_aux.STATES[2]:                                      # state3 or surplus
            possibles_modes.append(fct_aux.STATE3_STRATS)
            # print("3: num_pl_i={}, state_i = {}".format(num_pl_i, state_i))
        
    return possibles_modes
# __________            find possibles modes --> debut               _________

# __________                compute prices --> debut                  _________
def compute_ben_cst_players(arr_pl_M_t_vars_mode_prof, gamma_is, b0_t, c0_t):
    """
    calculate the benefit and the cost of each player at time t

    Parameters
    ----------
    arr_pl_M_t_vars_mode_prof : array of shape M_PLAYERS, len(AUTOMATE_INDEX_ATTRS)
        DESCRIPTION.
    gamma_is :  array of shape (M_PLAYERS,)
        DESCRIPTION.
    
    b0_t : float
        benefit per unit.
    c0_t : float
        cost per unit.

    Returns
    -------
    bens: benefits of M_PLAYERS, shape (M_PLAYERS,).
    csts: costs of M_PLAYERS, shape (M_PLAYERS,)
    """
    bens_t = b0_t \
            * arr_pl_M_t_vars_mode_prof[:, 
                                        fct_aux.AUTOMATE_INDEX_ATTRS["prod_i"]] \
            + gamma_is \
                * arr_pl_M_t_vars_mode_prof[:, 
                                            fct_aux.AUTOMATE_INDEX_ATTRS["r_i"]]
    csts_t = c0_t \
            * arr_pl_M_t_vars_mode_prof[:, 
                                        fct_aux.AUTOMATE_INDEX_ATTRS["cons_i"]]
    return bens_t, csts_t

def compute_prices_inside_SG(arr_pl_M_t_vars_mode_prof,
                             sum_diff_pos_minus_0_t_minus_2,
                             sum_diff_pos_plus_0_t_minus_2,
                             sum_cons_is_0_t_minus_2,                             
                             sum_prod_is_0_t_minus_2,
                             pi_hp_plus, pi_hp_minus,
                             a, b,
                             pi_0_plus_t, pi_0_minus_t,
                             manual_debug, dbg):
    
    In_sg = np.sum(arr_pl_M_t_vars_mode_prof[
                            :,
                            fct_aux.AUTOMATE_INDEX_ATTRS["prod_i"]])
    Out_sg = np.sum(arr_pl_M_t_vars_mode_prof[
                        :,
                        fct_aux.AUTOMATE_INDEX_ATTRS["cons_i"]])
    sum_cons_is_0_t_minus_1 = sum_cons_is_0_t_minus_2 + Out_sg
    sum_prod_is_0_t_minus_1 = sum_prod_is_0_t_minus_2 + In_sg
    sum_diff_pos_minus_t = fct_aux.fct_positive(sum_list1=Out_sg, 
                                                sum_list2=In_sg)
    sum_diff_pos_plus_t = fct_aux.fct_positive(sum_list1=In_sg, 
                                               sum_list2=Out_sg)
    sum_diff_pos_minus_0_t_minus_1 = sum_diff_pos_minus_0_t_minus_2 \
                                     + sum_diff_pos_minus_t
    sum_diff_pos_plus_0_t_minus_1 = sum_diff_pos_plus_0_t_minus_2 \
                                        + sum_diff_pos_plus_t
    
    # compute the new prices pi_sg_plus_t, pi_sg_minus_t
    # from a pricing model in the document
    phi_hp_minus = fct_aux.compute_cost_energy_bought_by_SG_2_HP(
                        pi_hp_minus=pi_hp_minus, 
                        quantity=sum_diff_pos_minus_0_t_minus_1, b=b)
    phi_hp_plus = fct_aux.compute_benefit_energy_sold_by_SG_2_HP(
                        pi_hp_plus=pi_hp_plus, 
                        quantity=sum_diff_pos_plus_0_t_minus_1, a=a)
    pi_sg_minus_t = phi_hp_minus / sum_cons_is_0_t_minus_1 \
                        if sum_cons_is_0_t_minus_1 != 0 \
                        else 0
    pi_sg_plus_t = phi_hp_plus / sum_prod_is_0_t_minus_1 \
                        if sum_prod_is_0_t_minus_1 != 0 \
                        else 0
    
    ## compute prices inside smart grids
    # compute prices of an energy unit price for cost and benefit players
    b0_t, c0_t = fct_aux.compute_energy_unit_price(
                    pi_0_plus_t, pi_0_minus_t, 
                    pi_hp_plus, pi_hp_minus,
                    a, b,
                    In_sg, Out_sg)
    
    # compute ben, cst of shape (M_PLAYERS,) 
    # compute cost (csts) and benefit (bens) players by energy exchanged.
    gamma_is = arr_pl_M_t_vars_mode_prof[:, 
                                         fct_aux.AUTOMATE_INDEX_ATTRS["gamma_i"]]
    bens_t, csts_t = compute_ben_cst_players(arr_pl_M_t_vars_mode_prof, 
                                             gamma_is, b0_t, c0_t)
    
    return In_sg, Out_sg, \
            b0_t, c0_t, \
            bens_t, csts_t, \
            pi_sg_plus_t, pi_sg_minus_t
    
# __________                compute prices --> fin                  _________

# ________       balanced players 4 one modes_profile   ---> debut      ______
def balanced_player_game_4_mode_profil(arr_pl_M_t_vars_mode_prof, 
                                       m_players,
                                       dbg):
    """
    attribute modes of all players and get players' variables as prod_i, 
    cons_i, r_i, gamma_i saved to  arr_pl_M_T_vars_mode_prof

    Parameters
    ----------
    arr_pl_M_t_vars_mode_prof : shape (m_players, len(AUTOMATE_INDEX_ATTRS))
        DESCRIPTION.
    m_players : number of players
        DESCRIPTION.

    Returns
    -------
    arr_pl_M_t_vars_mode_prof

    """
    
    for num_pl_i in range(0, m_players):
        Pi = arr_pl_M_t_vars_mode_prof[num_pl_i, 
                                       fct_aux.AUTOMATE_INDEX_ATTRS['Pi']]
        Ci = arr_pl_M_t_vars_mode_prof[num_pl_i,
                                       fct_aux.AUTOMATE_INDEX_ATTRS['Ci']]
        Si = arr_pl_M_t_vars_mode_prof[num_pl_i, 
                                       fct_aux.AUTOMATE_INDEX_ATTRS['Si']]
        Si_max = arr_pl_M_t_vars_mode_prof[num_pl_i,
                                        fct_aux.AUTOMATE_INDEX_ATTRS['Si_max']]
        gamma_i = arr_pl_M_t_vars_mode_prof[num_pl_i,
                                       fct_aux.AUTOMATE_INDEX_ATTRS['gamma_i']]
        state_i = arr_pl_M_t_vars_mode_prof[num_pl_i,
                                 fct_aux.AUTOMATE_INDEX_ATTRS['state_i']]
        mode_i = arr_pl_M_t_vars_mode_prof[num_pl_i,
                                 fct_aux.AUTOMATE_INDEX_ATTRS['mode_i']]
        
        prod_i, cons_i, r_i = 0, 0, 0
        pl_i = players.Player(Pi, Ci, Si, Si_max, gamma_i, 
                              prod_i, cons_i, r_i, state_i)
        
        pl_i.set_R_i_old(Si_max-Si)
        pl_i.set_mode_i(mode_i)
        
        # update prod, cons and r_i
        pl_i.update_prod_cons_r_i()
        
        # is pl_i balanced?
        boolean, formule = fct_aux.balanced_player(pl_i, thres=0.1)
        
        # update variables in arr_pl_M_T_modif
        tup_cols_values = [("prod_i", pl_i.get_prod_i()), 
                ("cons_i", pl_i.get_cons_i()), ("r_i", pl_i.get_r_i()),
                ("R_i_old", pl_i.get_R_i_old()), ("Si", pl_i.get_Si()),
                ("Si_old", pl_i.get_Si_old()), ("mode_i", mode_i), 
                ("gamma_i", gamma_i),
                ("balanced_pl_i", boolean), ("formule", formule)]
        for col, val in tup_cols_values:
            arr_pl_M_t_vars_mode_prof[num_pl_i, 
                                    fct_aux.AUTOMATE_INDEX_ATTRS[col]] = val
            
    return arr_pl_M_t_vars_mode_prof

# ________       balanced players 4 one modes_profile   --->   fin      ______


# ________       balanced players 4 all modes_profiles   ---> debut      ______
def generer_balanced_players_4_modes_profils(arr_pl_M_t_vars_modif, 
                                             m_players, t,
                                             sum_diff_pos_minus_0_t_minus_2,
                                             sum_diff_pos_plus_0_t_minus_2,
                                             sum_cons_is_0_t_minus_2,                             
                                             sum_prod_is_0_t_minus_2,
                                             pi_hp_plus, pi_hp_minus,
                                             a, b,
                                             pi_0_plus_t, pi_0_minus_t,
                                             manual_debug, dbg):
    """
    generate the combinaison of all modes' profils and 
    for each modes' profil, balance the players' game
    
    parameters:
        
    sum_diff_pos_minus_0_t_minus_2 : sum of the positive difference btw cons_is and prod_is from 0 to t-2
        sum_{k=0}^{t-2} (|sum_{i=1}^{M} cons_i^k - sum_{i=1}^{M} prod_i^k|)
    sum_diff_pos_plus_0_t_minus_2 : sum of the positive difference btw prod_is and cons_is from 0 to t-2
        sum_{k=0}^{t-2} (|sum_{i=1}^{M} cons_i^k - sum_{i=1}^{M} prod_i^k|)
    sum_cons_is_0_t_minus_2 : sum of the cons of all players from 0 to t-2 (t-1 periods)
        sum_{k=0}^{t-2} (|sum_{i=1}^{M} cons_i^k)
    sum_prod_is_0_t_minus_2 : sum of the prod of all players from 0 to t-2 (t-1 periods)
        sum_{k=0}^{t-2} (|sum_{i=1}^{M} prod_i^k
    
    M = m_players
    
    """
    best_key_Perf_t, mid_key_Perf_t, bad_key_Perf_t = None, None, None
    dico_modes_profs_by_players_t_best = dict() 
    dico_modes_profs_by_players_t_bad = dict()
    dico_modes_profs_by_players_t_mid = dict()
    possibles_modes = possibles_modes_players_automate(
                                        arr_pl_M_t_vars_modif, m_players)
    print("possibles_modes={}".format(len(possibles_modes)))
    mode_profiles = it.product(*possibles_modes)
        
    Perf_ts = list()
    cpt_xxx = 0
    for mode_profile in mode_profiles:
        arr_pl_M_t_vars_mode_prof = arr_pl_M_t_vars_modif.copy()
        arr_pl_M_t_vars_mode_prof[:, 
                                  fct_aux.AUTOMATE_INDEX_ATTRS["mode_i"]] \
            = mode_profile
        
        arr_pl_M_t_vars_mode_prof = balanced_player_game_4_mode_profil(
                                       arr_pl_M_t_vars_mode_prof, 
                                       m_players,
                                       dbg)
        
        # compute pi_sg_{plus,minus}_t, pi_0_{plus,minus}_t, b0_t, c0_t, ben_t, cst_t
        In_sg, Out_sg, b0_t, c0_t = None, None, None, None 
        bens_t, csts_t = None, None 
        pi_sg_plus_t, pi_sg_minus_t = None, None
        In_sg, Out_sg, \
        b0_t, c0_t, \
        bens_t, csts_t, \
        pi_sg_plus_t, pi_sg_minus_t \
            = compute_prices_inside_SG(arr_pl_M_t_vars_mode_prof, 
                                       sum_diff_pos_minus_0_t_minus_2,
                                       sum_diff_pos_plus_0_t_minus_2,
                                       sum_cons_is_0_t_minus_2,                             
                                       sum_prod_is_0_t_minus_2,
                                       pi_hp_plus, pi_hp_minus,
                                       a, b,
                                       pi_0_plus_t, pi_0_minus_t,
                                       manual_debug, dbg)
        
        # compute Perf_t
        bens_csts_t = bens_t - csts_t
        Perf_t = np.sum(bens_csts_t, axis=0)
        
        dico_mode_prof_by_players = dict()
        for num_pl_i in range(0, m_players):
            dico_vars = dict()
            Vi = bens_csts_t[num_pl_i]
            dico_vars["Vi"] = round(Vi, 2)
            dico_vars["ben_i"] = round(bens_t[num_pl_i], 2)
            dico_vars["cst_i"] = round(csts_t[num_pl_i], 2)
            variables = ["set", "state_i", "mode_i", "Pi", "Ci", "Si_max", 
                         "Si_old", "Si", "prod_i", "cons_i", "r_i", 
                         "Si_minus", "Si_plus", "gamma_i"]
            for variable in variables:
                dico_vars[variable] = arr_pl_M_t_vars_mode_prof[
                                        num_pl_i, 
                                        fct_aux.AUTOMATE_INDEX_ATTRS[variable]]
                
            dico_mode_prof_by_players[fct_aux.RACINE_PLAYER
                                        +str(num_pl_i)
                                        +"_t_"+str(t)
                                        +"_"+str(cpt_xxx)] \
                = dico_vars
        
        dico_mode_prof_by_players["bens_t"] = bens_t
        dico_mode_prof_by_players["csts_t"] = csts_t
        dico_mode_prof_by_players["Perf_t"] = round(Perf_t, 2)                  # utility of the game
        dico_mode_prof_by_players["b0_t"] = round(b0_t,2)
        dico_mode_prof_by_players["c0_t"] = round(c0_t,2)
        dico_mode_prof_by_players["Out_sg"] = round(Out_sg,2)
        dico_mode_prof_by_players["In_sg"] = round(In_sg,2)
        dico_mode_prof_by_players["pi_sg_plus_t"] = round(pi_sg_plus_t,2)
        dico_mode_prof_by_players["pi_sg_minus_t"] = round(pi_sg_minus_t,2)
        dico_mode_prof_by_players["pi_0_plus_t"] = round(pi_0_plus_t,2)
        dico_mode_prof_by_players["pi_0_minus_t"] = round(pi_0_minus_t,2)
        dico_mode_prof_by_players["mode_profile"] = mode_profile
            
        best_key_Perf_t = select_perf_t_from_key(best=best_key_Perf_t,
                                                 current=Perf_t,
                                                 key="best")
        bad_key_Perf_t = select_perf_t_from_key(best=bad_key_Perf_t,
                                                 current=Perf_t,
                                                 key="bad")
        Perf_ts.append(Perf_t)
        id_mid_key_Perf_t = np.argsort(Perf_ts)[len(Perf_ts)//2]
        mid_key_Perf_t_new = Perf_ts[id_mid_key_Perf_t]
            
        if best_key_Perf_t == Perf_t \
            and best_key_Perf_t not in dico_modes_profs_by_players_t_best:
            dico_modes_profs_by_players_t_best = dict()
            dico_modes_profs_by_players_t_best[Perf_t] \
                = [ ("BF_{}_t_{}".format(cpt_xxx,t), 
                      dico_mode_prof_by_players) ]
        elif best_key_Perf_t == Perf_t \
            and best_key_Perf_t in dico_modes_profs_by_players_t_best:
            dico_modes_profs_by_players_t_best[Perf_t]\
            .append( ("BF_{}_t_{}".format(cpt_xxx,t), 
                      dico_mode_prof_by_players) )
            
        if bad_key_Perf_t == Perf_t \
            and bad_key_Perf_t not in dico_modes_profs_by_players_t_bad:
            dico_modes_profs_by_players_t_bad = dict()
            dico_modes_profs_by_players_t_bad[Perf_t] \
                = [ ("BF_{}_t_{}".format(cpt_xxx,t), 
                      dico_mode_prof_by_players) ]
        elif bad_key_Perf_t == Perf_t \
            and bad_key_Perf_t in dico_modes_profs_by_players_t_bad:
            dico_modes_profs_by_players_t_bad[Perf_t]\
            .append( ("BF_{}_t_{}".format(cpt_xxx,t), 
                      dico_mode_prof_by_players) )
        
        mid_key_Perf_t_OLD = mid_key_Perf_t
        if mid_key_Perf_t != mid_key_Perf_t_new:
            dico_modes_profs_by_players_t_mid = dict()
            dico_modes_profs_by_players_t_mid[mid_key_Perf_t_new] \
                = [ ("BF_{}_t_{}".format(cpt_xxx,t), 
                      dico_mode_prof_by_players) ]
            mid_key_Perf_t = mid_key_Perf_t_new
            
        # print("Perf_t={}, dico_best={}, dico_bad={} mid_key_Perf_t_OLD={}, mid_key_Perf_t_new={},dico_mid={}".format(
        #      Perf_t, list(dico_modes_profs_by_players_t_best.keys()), 
        #      list( dico_modes_profs_by_players_t_bad.keys() ),
        #      mid_key_Perf_t_OLD, mid_key_Perf_t_new,
        #      list( dico_modes_profs_by_players_t_mid.keys() )
        #      ))
            
        
        cpt_xxx += 1
        print("cpt_xxx={}, After running free memory={}%".format(cpt_xxx,
                    list(psutil.virtual_memory())[2]    )) \
            if cpt_xxx % MOD ==0 else None
        
   
    print("Perf_t: BAD={}, MIDDLE={}, BEST={}".format(bad_key_Perf_t, mid_key_Perf_t, best_key_Perf_t))
        
    list_dico_modes_profs_by_players_t_best \
        = dico_modes_profs_by_players_t_best[best_key_Perf_t]
    list_dico_modes_profs_by_players_t_bad \
        = dico_modes_profs_by_players_t_bad[bad_key_Perf_t]
    list_dico_modes_profs_by_players_t_mid \
        = dico_modes_profs_by_players_t_mid[mid_key_Perf_t]
      
    return list_dico_modes_profs_by_players_t_best, \
            list_dico_modes_profs_by_players_t_bad, \
            list_dico_modes_profs_by_players_t_mid
# _______       balanced players 4 all modes_profils   --->  fin       _______

# _______           sum of prod, cons from 0 to t-2 ---> Debut        _________
def get_sum_cons_prod_from_0_t_minus_2(arr_pl_M_T_vars_modif, t):
    """
    quantity of energies (prod_is, cons_is) from 0 to t-2 to get values 
    for t-1 periods
    """ 
    if t == 0:
        sum_diff_pos_minus_0_t_minus_2 = 0
        sum_diff_pos_plus_0_t_minus_2 = 0
        sum_cons_is_0_t_minus_2 = 0                                        
        sum_prod_is_0_t_minus_2 = 0
    else:
        # TODO : utiliser les vecteurs pour calculer la somme
        # compute the positive difference btw cons_is and prod_is from 0 to t-1:
            # sum_diff_pos_minus_0_t_minus_1
        # compute the positive difference btw prod_is and cons_is from 0 to t-1
            # sum_diff_pos_plus_0_t_minus_1
        sum_cons_is_0_t_minus_2 \
            = np.sum( np.sum(arr_pl_M_T_vars_modif[
                                :,0:t,
                                fct_aux.AUTOMATE_INDEX_ATTRS["cons_i"]], 
                             axis=0), 
                     axis=0)
        sum_prod_is_0_t_minus_2 \
            = np.sum( np.sum(arr_pl_M_T_vars_modif[
                                :,0:t,
                                fct_aux.AUTOMATE_INDEX_ATTRS["prod_i"]], 
                             axis=0), 
                     axis=0)
        
        sum_diff_pos_minus_0_t_minus_2 = 0
        sum_diff_pos_plus_0_t_minus_2 = 0
        for k in range(0,t):
            cons_k_is = np.sum(arr_pl_M_T_vars_modif[
                                :,k, 
                                fct_aux.AUTOMATE_INDEX_ATTRS["cons_i"]], 
                            axis=0)
            prod_k_is = np.sum(arr_pl_M_T_vars_modif[
                                :,k, 
                                fct_aux.AUTOMATE_INDEX_ATTRS["prod_i"]], 
                            axis=0)
            diff_pos_minus_k = fct_aux.fct_positive(sum_list1=cons_k_is, 
                                                    sum_list2=prod_k_is)
            diff_pos_plus_k = fct_aux.fct_positive(sum_list1=prod_k_is, 
                                                    sum_list2=cons_k_is)
            sum_diff_pos_minus_0_t_minus_2 += diff_pos_minus_k
            sum_diff_pos_plus_0_t_minus_2 += diff_pos_plus_k 
            
    return sum_diff_pos_minus_0_t_minus_2, \
            sum_diff_pos_plus_0_t_minus_2, \
            sum_cons_is_0_t_minus_2, \
            sum_prod_is_0_t_minus_2
# _______           sum of prod, cons from 0 to t-2 ---> fin        _________

# ____________        checkout NASH equilibrium --> debut        ______________
def checkout_nash_4_profils_by_periods(arr_pl_M_t_vars_modif_algo,
                                        arr_pl_M_t_vars_init,
                                        sum_diff_pos_minus_0_t_minus_2,
                                        sum_diff_pos_plus_0_t_minus_2,
                                        sum_cons_is_0_t_minus_2,                             
                                        sum_prod_is_0_t_minus_2,
                                        pi_hp_plus, pi_hp_minus, 
                                        a, b,
                                        pi_0_minus_t, pi_0_plus_t, 
                                        bens_csts_M_t,
                                        m_players,
                                        t,
                                        manual_debug,
                                        dbg):
    """
    verify if the modes' profil of players at time t is a Nash equilibrium.
    """
    # create a result dataframe of checking players' stability and nash equilibrium
    cols = ["players", "nash_modes_t{}".format(t), 'states_t{}'.format(t), 
            'Vis_t{}'.format(t), 'Vis_bar_t{}'.format(t), 
               'res_t{}'.format(t)] 
    
    id_players = list(range(0, m_players))
    df_nash_t = pd.DataFrame(index=id_players, columns=cols)
    
    # revert Si to the initial value ie at t and k=0
    Sis_init = arr_pl_M_t_vars_init[:, fct_aux.AUTOMATE_INDEX_ATTRS["Si"]]
    arr_pl_M_t_vars_modif_algo[:, fct_aux.AUTOMATE_INDEX_ATTRS["Si"]] = Sis_init
    
    # stability of each player
    modes_profil = list(arr_pl_M_t_vars_modif_algo[
                            :,
                            fct_aux.AUTOMATE_INDEX_ATTRS["mode_i"]] )
    for num_pl_i in range(0, m_players):
        state_i = arr_pl_M_t_vars_modif_algo[
                        num_pl_i,
                        fct_aux.AUTOMATE_INDEX_ATTRS["state_i"]] 
        mode_i = modes_profil[num_pl_i]
        mode_i_bar = fct_aux.find_out_opposite_mode(state_i, mode_i)
        
        opposite_modes_profil = modes_profil.copy()
        opposite_modes_profil[num_pl_i] = mode_i_bar
        opposite_modes_profil = tuple(opposite_modes_profil)
        
        df_nash_t.loc[num_pl_i, "players"] = fct_aux.RACINE_PLAYER+"_"+str(num_pl_i)
        df_nash_t.loc[num_pl_i, "nash_modes_t{}".format(t)] = mode_i
        df_nash_t.loc[num_pl_i, "states_t{}".format(t)] = state_i
        
        arr_pl_M_t_vars_modif_algo = balanced_player_game_4_mode_profil(
                                             arr_pl_M_t_vars_modif_algo, 
                                             m_players,
                                             dbg)
        ## test if there are the same values like these in dico_mode_prof_by_players
        In_sg_bar, Out_sg_bar, b0_t_bar, c0_t_bar = None, None, None, None
        bens_t_bar, csts_t_bar = None, None
        pi_sg_plus_t_bar, pi_sg_minus_t_bar = None, None
        In_sg_bar, Out_sg_bar, \
        b0_t_bar, c0_t_bar, \
        bens_t_bar, csts_t_bar, \
        pi_sg_plus_t_bar, pi_sg_minus_t_bar \
            = compute_prices_inside_SG(
                arr_pl_M_t_vars_modif_algo, 
                sum_diff_pos_minus_0_t_minus_2,
                sum_diff_pos_plus_0_t_minus_2,
                sum_cons_is_0_t_minus_2,                             
                sum_prod_is_0_t_minus_2,
                pi_hp_plus, pi_hp_minus,
                a, b,
                pi_0_plus_t, pi_0_minus_t,
                manual_debug, dbg)
        bens_csts_t_bar = bens_t_bar - csts_t_bar
        
        Vi = bens_csts_M_t[num_pl_i]
        Vi_bar = bens_csts_t_bar[num_pl_i]
        
        df_nash_t.loc[num_pl_i, 'Vis_t{}'.format(t)] = Vi
        df_nash_t.loc[num_pl_i, 'Vis_bar_t{}'.format(t)] = Vi_bar
        res = None
        if Vi >= Vi_bar:
            res = "STABLE"
            df_nash_t.loc[num_pl_i, 'res_t{}'.format(t)] = res
        else:
            res = "INSTABLE"
            df_nash_t.loc[num_pl_i, 'res_t{}'.format(t)] = res   
            
    return df_nash_t
    
# ____________        checkout NASH equilibrium --> fin          ______________

# ____________          turn dico in2 df  --> debut             ______________
def turn_dico_stats_res_into_df_BF(dico_modes_profs_players_algo, 
                                  path_to_save, 
                                  t_periods, 
                                  manual_debug, 
                                  algo_name):
    """
    transform the dico into a DataFrame

    """
    df = None
    for t in range(0, t_periods):
        dico_bf_t = dico_modes_profs_players_algo[t]
        dico_bf_t_new = dict()
        
        Perf_t = dico_bf_t["Perf_t"]
        b0_t = dico_bf_t["b0_t"] 
        c0_t = dico_bf_t["c0_t"]
        Out_sg = dico_bf_t["Out_sg"]
        In_sg = dico_bf_t["In_sg"]
        pi_sg_plus_t = dico_bf_t["pi_sg_plus_t"]
        pi_sg_minus_t = dico_bf_t["pi_sg_minus_t"]
        pi_0_plus_t = dico_bf_t["pi_0_plus_t"]
        pi_0_minus_t = dico_bf_t["pi_0_minus_t"]
        mode_profile = dico_bf_t["mode_profile"]
        bens_t = dico_bf_t["bens_t"]
        csts_t = dico_bf_t["csts_t"]
        dico_keys_2_delete = {'Out_sg':Out_sg, 'In_sg':In_sg, 
            'Perf_t':Perf_t, 'b0_t':b0_t, 'c0_t':c0_t,
            'pi_sg_plus_t':pi_sg_plus_t, 'pi_sg_minus_t':pi_sg_minus_t, 
            'pi_0_plus_t':pi_0_plus_t, 'pi_0_minus_t':pi_0_minus_t, 
            'mode_profile': mode_profile, 
            'bens_t':bens_t, 'csts_t':csts_t}
        
        for key, val in dico_keys_2_delete.items():
            dico_bf_t.pop(key, None)
                
        keys_to_delete = ['mode_profile', 'bens_t', 'csts_t']
        dico_vals_new = dict()
        for player, dico_vals in dico_bf_t.items():
            if player not in dico_keys_2_delete.keys():
                cpt_combi = player.split('_')[3]
                player = player.split('_')[0]
                dico_vals_new["cpt_comb_"+str(t)] = cpt_combi
                for key, val in dico_keys_2_delete.items():
                    dico_vals.pop(key, None)
                    if key not in keys_to_delete:
                        #dico_vals[key "{}_t{}".format(key,t)] = val
                        dico_vals_new["{}_t{}".format(key,t)] = val
            for k, v in dico_vals.items():
                dico_vals_new["{}_t{}".format(k,t)] = v
            dico_bf_t_new[player] = dico_vals_new
                
        # df_t = pd.DataFrame.from_dict(dico_bf_t_new, orient='columns')
        df_t = pd.DataFrame.from_dict(dico_bf_t_new)
        if df is None:
            df = df_t.copy()
        else:
            df = pd.concat([df, df_t], axis=0)
                
    # save df to xlsx
    df.to_excel(os.path.join(*[path_to_save,
                               "{}_dico_BF.xlsx".format(algo_name)]), 
                index=True)
    
def turn_dico_stats_res_into_df_BF_FAUX(dico_modes_profs_players_algo, 
                                  path_to_save, 
                                  t_periods, 
                                  manual_debug, 
                                  algo_name):
    """
    transform the dico into a DataFrame

    """
    df = None
    for t in range(0, t_periods):
        dico_bf_t = dico_modes_profs_players_algo[t]
        dico_bf_t_new = dict()
        
        Perf_t = dico_bf_t["Perf_t"]
        b0_t = dico_bf_t["b0_t"] 
        c0_t = dico_bf_t["c0_t"]
        Out_sg = dico_bf_t["Out_sg"]
        In_sg = dico_bf_t["In_sg"]
        pi_sg_plus_t = dico_bf_t["pi_sg_plus_t"]
        pi_sg_minus_t = dico_bf_t["pi_sg_minus_t"]
        pi_0_plus_t = dico_bf_t["pi_0_plus_t"]
        pi_0_minus_t = dico_bf_t["pi_0_minus_t"]
        mode_profile = dico_bf_t["mode_profile"]
        bens_t = dico_bf_t["bens_t"]
        csts_t = dico_bf_t["csts_t"]
        dico_keys_2_delete = {'Out_sg':Out_sg, 'In_sg':In_sg, 
            'Perf_t':Perf_t, 'b0_t':b0_t, 'c0_t':c0_t,
            'pi_sg_plus_t':pi_sg_plus_t, 'pi_sg_minus_t':pi_sg_minus_t, 
            'pi_0_plus_t':pi_0_plus_t, 'pi_0_minus_t':pi_0_minus_t}
        
        keys_to_delete = ['mode_profile', 'bens_t', 'csts_t']
        
        for player, dico_vals in dico_bf_t.items():
            dico_tmp = dict()
            for key, val in dico_vals.items():
                dico_tmp["{}_t{}".format(key,t)] = val
            if player not in dico_keys_2_delete.keys() \
                and player not in keys_to_delete:
                player = player.split('_')[0]
                cpt_combi = player.split('_')[3]
                dico_tmp["cpt_comb_"+str(t)] = cpt_combi
                for k, v in dico_keys_2_delete.items():
                    dico_tmp["{}_t{}".format(k,t)] = v
                    
            dico_bf_t_new[player] = dico_tmp
        
        # df_t = pd.DataFrame.from_dict(dico_bf_t_new, orient='columns')
        df_t = pd.DataFrame.from_dict(dico_bf_t_new)
        if df is None:
            df = df_t.copy()
        else:
            df = pd.concat([df, df_t], axis=0)
                
    # save df to xlsx
    df.to_excel(os.path.join(*[path_to_save,
                               "{}_dico_BF.xlsx".format(algo_name)]), 
                index=True)

# ____________          turn dico in2 df  -->   fin             ______________

# _________          add gamma & state 4 players  -->   debut         _________
def get_values_Pi_Ci_Si_Simax_Pi1_Ci1(arr_pl_M_t_K_vars, 
                                      arr_pl_M_t_minus_1_K_vars,
                                      arr_pl_M_t_plus_1_K_vars,
                                      num_pl_i, k,
                                      shape_arr_pl):
    """
    return the values of Pi, Ci, Si, Si_max, Pi_t_plus_1, Ci_t_plus_1 from arrays
    """
    Pi = arr_pl_M_t_K_vars[num_pl_i, fct_aux.AUTOMATE_INDEX_ATTRS["Pi"]] \
        if shape_arr_pl == 2 \
        else arr_pl_M_t_K_vars[num_pl_i, k, 
                               fct_aux.AUTOMATE_INDEX_ATTRS["Pi"]]
    Ci = arr_pl_M_t_K_vars[num_pl_i, fct_aux.AUTOMATE_INDEX_ATTRS["Ci"]] \
        if shape_arr_pl == 2 \
        else arr_pl_M_t_K_vars[num_pl_i, k, 
                               fct_aux.AUTOMATE_INDEX_ATTRS["Pi"]]
    Si_max = arr_pl_M_t_K_vars[num_pl_i, 
                               fct_aux.AUTOMATE_INDEX_ATTRS["Si_max"]] \
        if shape_arr_pl == 2 \
        else arr_pl_M_t_K_vars[num_pl_i, k, 
                               fct_aux.AUTOMATE_INDEX_ATTRS["Si_max"]]
    Si = arr_pl_M_t_minus_1_K_vars[num_pl_i, 
                               fct_aux.AUTOMATE_INDEX_ATTRS["Si"]] \
        if shape_arr_pl == 2 \
        else arr_pl_M_t_minus_1_K_vars[num_pl_i, k,
                               fct_aux.AUTOMATE_INDEX_ATTRS["Si"]]
    Ci_t_plus_1 = arr_pl_M_t_plus_1_K_vars[num_pl_i,
                                   fct_aux.AUTOMATE_INDEX_ATTRS["Ci"]] \
        if shape_arr_pl == 2 \
        else arr_pl_M_t_plus_1_K_vars[num_pl_i,
                                   fct_aux.AUTOMATE_INDEX_ATTRS["Ci"]]
    Pi_t_plus_1 = arr_pl_M_t_plus_1_K_vars[num_pl_i,
                                   fct_aux.AUTOMATE_INDEX_ATTRS["Pi"]] \
        if shape_arr_pl == 2 \
        else arr_pl_M_t_plus_1_K_vars[num_pl_i,
                                       fct_aux.AUTOMATE_INDEX_ATTRS["Pi"]]
    
    return Pi, Ci, Si, Si_max, Pi_t_plus_1, Ci_t_plus_1

def update_variables_MtK(arr_pl_M_t_K_vars, arr_pl_M_t_minus_1_K_vars, 
                         variables, shape_arr_pl,
                         num_pl_i, t, k, gamma_i, Si,
                         pi_0_minus, pi_0_plus, 
                         pi_hp_minus_t, pi_hp_plus_t, dbg):
    # ____              update cell arrays: debut               _______
    if shape_arr_pl == 3:
        for (var,val) in variables:
            arr_pl_M_t_K_vars[num_pl_i, :,
                        fct_aux.AUTOMATE_INDEX_ATTRS[var]] = val
            
    elif shape_arr_pl == 2:
        for (var,val) in variables:
            arr_pl_M_t_K_vars[num_pl_i,
                    fct_aux.AUTOMATE_INDEX_ATTRS[var]] = val
    # ____              update cell arrays: fin                 _______
    
    bool_gamma_i = (gamma_i >= min(pi_0_minus, pi_0_plus)-1) \
                    & (gamma_i <= max(pi_hp_minus_t, pi_hp_plus_t)+1)
    print("GAMMA : player={}, val={}, bool_gamma_i={}"\
          .format(num_pl_i, gamma_i, bool_gamma_i)) if dbg else None

    arr_pl_M_t_minus_1_K_vars[num_pl_i, 
                                       fct_aux.AUTOMATE_INDEX_ATTRS["Si"]] \
                            if shape_arr_pl == 2 \
                            else arr_pl_M_t_minus_1_K_vars[num_pl_i, k, 
                                             fct_aux.AUTOMATE_INDEX_ATTRS["Si"]]

    Si_t_minus_1 = None
    if shape_arr_pl == 2 and t > 0:
        Si_t_minus_1 = arr_pl_M_t_minus_1_K_vars[num_pl_i, 
                                       fct_aux.AUTOMATE_INDEX_ATTRS["Si"]] 
    elif shape_arr_pl == 2 and t == 0:
        Si_t_minus_1 = arr_pl_M_t_K_vars[num_pl_i, 
                               fct_aux.AUTOMATE_INDEX_ATTRS["Si"]]
    elif shape_arr_pl == 3 and t > 0:
        Si_t_minus_1 = arr_pl_M_t_minus_1_K_vars[num_pl_i, k, 
                               fct_aux.AUTOMATE_INDEX_ATTRS["Si"]] 
    elif shape_arr_pl == 3 and t == 0:
        Si_t_minus_1 = arr_pl_M_t_K_vars[num_pl_i, k,
                               fct_aux.AUTOMATE_INDEX_ATTRS["Si"]]
        
    print("Si_t_minus_1={}, Si={}".format(Si_t_minus_1, Si)) \
        if dbg else None
            
    return arr_pl_M_t_K_vars

def compute_gamma_state_4_period_t(arr_pl_M_t_K_vars, 
                                   arr_pl_M_t_minus_1_K_vars,
                                   arr_pl_M_t_plus_1_K_vars,
                                   t,
                                   pi_0_plus, pi_0_minus,
                                   pi_hp_plus_t, pi_hp_minus_t,
                                   m_players,
                                   t_periods,
                                   gamma_version,
                                   manual_debug=False, dbg=False):
    """
    compute gamma_i et determinate the state for all players 
    
    arr_pl_M_T_K_vars: shape (m_players, len(vars)) or 
                             (m_players, k_steps, len(vars))
    """
    k = 0
    shape_arr_pl = len(arr_pl_M_t_K_vars.shape)
    
    # compute Cis_t_plus_t, Pis_t_plus_1, 
    # GC_t, GSis_t_minus, GSis_t_plus, Xis, Yis
    Cis_t_plus_1 = arr_pl_M_t_plus_1_K_vars[:,fct_aux.AUTOMATE_INDEX_ATTRS["Ci"]]
    Pis_t_plus_1 = arr_pl_M_t_plus_1_K_vars[:,fct_aux.AUTOMATE_INDEX_ATTRS["Pi"]]
    Cis_Pis_t_plus_1 = Cis_t_plus_1 - Pis_t_plus_1
    Cis_Pis_t_plus_1[Cis_Pis_t_plus_1 < 0] = 0
    GC_t = np.sum(Cis_Pis_t_plus_1)
    
    # initialisation of variables for gamma_version = 2
    state_is = np.empty(shape=(m_players,), dtype=object)
    Sis = np.zeros(shape=(m_players,))
    GSis_t_minus = np.zeros(shape=(m_players,)) 
    GSis_t_plus = np.zeros(shape=(m_players,))
    Xis = np.zeros(shape=(m_players,)) 
    Yis = np.zeros(shape=(m_players,))
    
    for num_pl_i in range(0, m_players):
        Pi, Ci, Si, Si_max, Pi_t_plus_1, Ci_t_plus_1 \
            = get_values_Pi_Ci_Si_Simax_Pi1_Ci1(
                arr_pl_M_t_K_vars, 
                arr_pl_M_t_minus_1_K_vars,
                arr_pl_M_t_plus_1_K_vars,
                num_pl_i, k,
                shape_arr_pl)
        
        prod_i, cons_i, r_i, gamma_i, state_i = 0, 0, 0, 0, ""
        pl_i = None
        pl_i = players.Player(Pi, Ci, Si, Si_max, gamma_i, 
                              prod_i, cons_i, r_i, state_i)
        state_i = pl_i.find_out_state_i()
        state_is[num_pl_i] = state_i
        
        Si_t_minus, Si_t_plus = None, None
        Xi, Yi, X_gamV5 = None, None, None
        if state_i == fct_aux.STATES[0]:                                       # state1 or Deficit
            Si_t_minus = 0
            Si_t_plus = Si
            Xi = pi_0_minus
            Yi = pi_hp_minus_t
            X_gamV5 = pi_0_minus
        elif state_i == fct_aux.STATES[1]:                                     # state2 or Self
            Si_t_minus = Si - (Ci - Pi)
            Si_t_plus = Si
            Xi = pi_0_minus
            Yi = pi_hp_minus_t
            X_gamV5 = pi_0_minus
        elif state_i == fct_aux.STATES[2]:                                     # state3 or Surplus
            Si_t_minus = Si
            Si_t_plus = max(Si_max, Si+(Pi-Ci))
            Xi = pi_0_plus
            Yi = pi_hp_plus_t
            X_gamV5 = pi_0_plus
        Sis[num_pl_i] = Si
        GSis_t_minus[num_pl_i] = Si_t_minus
        GSis_t_plus[num_pl_i] = Si_t_plus
        Xis[num_pl_i] = Xi; Yis[num_pl_i] = Yi
        
        gamma_i, gamma_i_min, gamma_i_max, gamma_i_mid  = None, None, None, None
        res_mid = None
        ppi_t = None
        if manual_debug:
            gamma_i_min = fct_aux.MANUEL_DBG_GAMMA_I
            gamma_i_mid = fct_aux.MANUEL_DBG_GAMMA_I
            gamma_i_max = fct_aux.MANUEL_DBG_GAMMA_I
            gamma_i = fct_aux.MANUEL_DBG_GAMMA_I
        else:
            Si_t_plus_1 = fct_aux.fct_positive(Ci_t_plus_1, Pi_t_plus_1)
            gamma_i_min = Xi - 1
            gamma_i_max = Yi + 1
            # print("Pi={}, Ci={}, Si={}, Si_t_plus_1={}, Si_t_minus={}, Si_t_plus={}".format(Pi, 
            #         Ci, Si, Si_t_plus_1, Si_t_minus, Si_t_plus))
            dif_pos_Ci_Pi_t_plus_1_Si_t_minus = 0
            dif_pos_Ci_Pi_t_plus_1_Si_t_minus = fct_aux.fct_positive(Si_t_plus_1, 
                                                                     Si_t_minus)
            dif_Si_plus_minus = Si_t_plus - Si_t_minus
            frac = dif_pos_Ci_Pi_t_plus_1_Si_t_minus / dif_Si_plus_minus \
                    if dif_Si_plus_minus != 0 else 0
            ppi_t = np.sqrt(min(frac, 1))
            
            if Si_t_plus_1 < Si_t_minus:
                # Xi - 1
                gamma_i = gamma_i_min
            elif Si_t_plus_1 >= Si_t_plus:
                # Yi + 1
                gamma_i = gamma_i_max
            elif Si_t_plus_1 >= Si_t_minus and Si_t_plus_1 < Si_t_plus:
                res_mid = ( Si_t_plus_1 - Si_t_minus) / \
                        (Si_t_plus - Si_t_minus)
                Z = Xi + (Yi-Xi)*res_mid
                gamma_i_mid = int(np.floor(Z))
                gamma_i = gamma_i_mid
                      
        
        if gamma_version == 0:
            gamma_i = 0
            variables = [("Si", Si), ("state_i", state_i), ("gamma_i", gamma_i), 
                      ("Si_minus", Si_t_minus), ("Si_plus", Si_t_plus)]
            arr_pl_M_t_K_vars = update_variables_MtK(
                                arr_pl_M_t_K_vars, arr_pl_M_t_minus_1_K_vars, 
                                variables, shape_arr_pl,
                                num_pl_i, t, k, gamma_i, Si,
                                pi_0_minus, pi_0_plus, 
                                pi_hp_minus_t, pi_hp_plus_t, dbg)
            
        if gamma_version == 1:
            variables = [("Si", Si), ("state_i", state_i), ("gamma_i", gamma_i), 
                      ("Si_minus", Si_t_minus), ("Si_plus", Si_t_plus)]
            arr_pl_M_t_K_vars = update_variables_MtK(
                                arr_pl_M_t_K_vars, arr_pl_M_t_minus_1_K_vars, 
                                variables, shape_arr_pl,
                                num_pl_i, t, k, gamma_i, Si,
                                pi_0_minus, pi_0_plus, 
                                pi_hp_minus_t, pi_hp_plus_t, dbg)
            
        elif gamma_version == 3:
            gamma_i = None
            if manual_debug:
                gamma_i = fct_aux.MANUEL_DBG_GAMMA_I
            elif Si_t_plus_1 < Si_t_minus:
                gamma_i = gamma_i_min
            else :
                gamma_i = gamma_i_max
            variables = [("Si", Si), ("state_i", state_i), ("gamma_i", gamma_i), 
                      ("Si_minus", Si_t_minus), ("Si_plus", Si_t_plus)]
            arr_pl_M_t_K_vars = update_variables_MtK(
                                arr_pl_M_t_K_vars, arr_pl_M_t_minus_1_K_vars, 
                                variables, shape_arr_pl,
                                num_pl_i, t, k, gamma_i, Si,
                                pi_0_minus, pi_0_plus, 
                                pi_hp_minus_t, pi_hp_plus_t, dbg)
            
        elif gamma_version == 4:
            gamma_i = None
            if manual_debug:
                gamma_i = fct_aux.MANUEL_DBG_GAMMA_I
            else:
                if Si_t_plus_1 < Si_t_minus:
                    # Xi - 1
                    gamma_i = gamma_i_min
                elif Si_t_plus_1 >= Si_t_plus:
                    # Yi + 1
                    gamma_i = gamma_i_max
                elif Si_t_plus_1 >= Si_t_minus and Si_t_plus_1 < Si_t_plus:
                    res_mid = ( Si_t_plus_1 - Si_t_minus) / \
                            (Si_t_plus - Si_t_minus)
                    Z = Xi + (Yi-Xi) * np.sqrt(res_mid)
                    gamma_i_mid = int(np.floor(Z))
                    gamma_i = gamma_i_mid
                
            variables = [("Si", Si), ("state_i", state_i), ("gamma_i", gamma_i), 
                     ("Si_minus", Si_t_minus), ("Si_plus", Si_t_plus)]
            arr_pl_M_t_K_vars = update_variables_MtK(
                                arr_pl_M_t_K_vars, arr_pl_M_t_minus_1_K_vars, 
                                variables, shape_arr_pl,
                                num_pl_i, t, k, gamma_i, Si,
                                pi_0_minus, pi_0_plus, 
                                pi_hp_minus_t, pi_hp_plus_t, dbg)
            
        elif gamma_version == -1:
            gamma_i = np.random.randint(low=2, high=21, size=1)[0]
            variables = [("Si", Si), ("state_i", state_i), ("gamma_i", gamma_i), 
                         ("Si_minus", Si_t_minus), ("Si_plus", Si_t_plus)]
            arr_pl_M_t_K_vars = update_variables_MtK(
                                arr_pl_M_t_K_vars, arr_pl_M_t_minus_1_K_vars, 
                                variables, shape_arr_pl,
                                num_pl_i, t, k, gamma_i, Si,
                                pi_0_minus, pi_0_plus, 
                                pi_hp_minus_t, pi_hp_plus_t, dbg)
                    
        elif gamma_version == 5:
            rd_draw = np.random.uniform(low=0.0, high=1.0, size=None)
            rho_i_t = 1 if rd_draw < ppi_t else 0 
            gamma_i = rho_i_t * (X_gamV5 + 1)
            variables = [("Si", Si), ("state_i", state_i), ("gamma_i", gamma_i), 
                         ("Si_minus", Si_t_minus), ("Si_plus", Si_t_plus)]
            arr_pl_M_t_K_vars = update_variables_MtK(
                                arr_pl_M_t_K_vars, arr_pl_M_t_minus_1_K_vars, 
                                variables, shape_arr_pl,
                                num_pl_i, t, k, gamma_i, Si,
                                pi_0_minus, pi_0_plus, 
                                pi_hp_minus_t, pi_hp_plus_t, dbg)
        elif gamma_version == -2:
            ppi_t = 0.8
            rd_draw = np.random.uniform(low=0.0, high=1.0, size=None)
            rho_i_t = 1 if rd_draw < ppi_t else 0 
            gamma_i = rho_i_t * (X_gamV5 + 1)
            variables = [("Si", Si), ("state_i", state_i), ("gamma_i", gamma_i), 
                         ("Si_minus", Si_t_minus), ("Si_plus", Si_t_plus)]
            arr_pl_M_t_K_vars = update_variables_MtK(
                                arr_pl_M_t_K_vars, arr_pl_M_t_minus_1_K_vars, 
                                variables, shape_arr_pl,
                                num_pl_i, t, k, gamma_i, Si,
                                pi_0_minus, pi_0_plus, 
                                pi_hp_minus_t, pi_hp_plus_t, dbg)
            
            
    if gamma_version == 2:
        GS_t_minus = np.sum(GSis_t_minus)
        GS_t_plus = np.sum(GSis_t_plus)
        gamma_is = None
        if GC_t <= GS_t_minus:
            gamma_is = Xis - 1
        elif GC_t > GS_t_plus:
            gamma_is = Yis + 1
        else:
            frac = (GC_t - GS_t_minus) / (GS_t_plus - GS_t_minus)
            res_is = Xis + (Yis-Xis)*frac
            gamma_is = np.floor(res_is)
            
        # ____              update cell arrays: debut               _______
        variables = [("Si", Sis), ("state_i", state_is), 
                     ("Si_minus", GSis_t_minus), ("Si_plus", GSis_t_plus)]
        if shape_arr_pl == 2:
            for (var,vals) in variables:
                arr_pl_M_t_K_vars[:, 
                        fct_aux.AUTOMATE_INDEX_ATTRS[var]] = vals
            if manual_debug:
                arr_pl_M_t_K_vars[:, 
                        fct_aux.AUTOMATE_INDEX_ATTRS["gamma_i"]] \
                    = fct_aux.MANUEL_DBG_GAMMA_I
            else:
                arr_pl_M_t_K_vars[:, 
                        fct_aux.AUTOMATE_INDEX_ATTRS["gamma_i"]] = gamma_is
        elif shape_arr_pl == 3:
            for (var,vals) in variables:
                arr_pl_M_t_K_vars[:, k,
                            fct_aux.AUTOMATE_INDEX_ATTRS[var]] = vals
                if manual_debug:
                    arr_pl_M_t_K_vars[:, k,
                            fct_aux.AUTOMATE_INDEX_ATTRS["gamma_i"]] \
                        = fct_aux.MANUEL_DBG_GAMMA_I
                else:
                    arr_pl_M_t_K_vars[:, 
                            fct_aux.AUTOMATE_INDEX_ATTRS["gamma_i"]] = gamma_is
                
        # ____              update cell arrays: fin               _______
        
        bool_gamma_is = (gamma_is >= min(pi_0_minus, pi_0_plus)-1) \
                            & (gamma_is <= max(pi_hp_minus_t, pi_hp_plus_t)+1)
        print("GAMMA : t={}, val={}, bool_gamma_is={}"\
              .format(t, gamma_is, bool_gamma_is)) if dbg else None
        GSis_t_minus_1 = arr_pl_M_t_minus_1_K_vars[
                            :, fct_aux.AUTOMATE_INDEX_ATTRS["Si"]] \
                        if shape_arr_pl == 2 \
                        else arr_pl_M_t_minus_1_K_vars[
                            :, k, fct_aux.AUTOMATE_INDEX_ATTRS["Si"]]
        print("GSis_t_minus_1={}, Sis={}".format(GSis_t_minus_1, Sis)) \
            if dbg else None
                
    return arr_pl_M_t_K_vars
# _________          add gamma & state 4 players  -->   fin           _________

    
    
###############################################################################
#                   definition  des fonctions principales
#
###############################################################################
# __________       main function of DETERMINIST   ---> debut      ____________
def bf_balanced_player_game(arr_pl_M_T_vars_init,
                            pi_hp_plus=0.02, 
                            pi_hp_minus=0.33,
                            a=1, b=1,
                            gamma_version=1,
                            path_to_save="tests", 
                            name_dir="tests", 
                            date_hhmm="DDMM_HHMM",
                            manual_debug=False, 
                            criteria_bf="Perf_t", dbg=False):
    """
    NE PLUS TOUCHER car TROP COMPLIQUE
    TODO GROS PROBLEME quand au regroupement des algos dans une meme fonction.
    En effet, les algos ont des etats differents et cela implique des stocks 
    differents. utiliser la fonction compute_gamma en dehors de la boucle for 
    des algos ne me permet pas de choisir le bon stock pour chaque algo.
    
    N'UTILISE CETTE FONCTION que pour t = 0
    pour t > 0 RESULTAT ERRONE
    """
    print("\n \n game: pi_hp_plus={}, pi_hp_minus={} ---> debut \n"\
          .format(pi_hp_plus, pi_hp_minus))
        
    m_players = arr_pl_M_T_vars_init.shape[0]
    t_periods = arr_pl_M_T_vars_init.shape[1]
    
    # _______ variables' initialization --> debut ________________
    pi_sg_plus_T = np.empty(shape=(t_periods,)); pi_sg_plus_T.fill(np.nan)
    pi_sg_minus_T = np.empty(shape=(t_periods,)); pi_sg_plus_T.fill(np.nan)
    pi_0_plus_T = np.empty(shape=(t_periods,)); pi_0_plus_T.fill(np.nan)
    pi_0_minus_T = np.empty(shape=(t_periods,)); pi_0_minus_T.fill(np.nan)
    B_is_M = np.empty(shape=(m_players,)); B_is_M.fill(np.nan)
    C_is_M = np.empty(shape=(m_players,)); C_is_M.fill(np.nan)
    B_is_M_T = np.empty(shape=(m_players, t_periods)); B_is_M_T.fill(np.nan)
    C_is_M_T = np.empty(shape=(m_players, t_periods)); C_is_M_T.fill(np.nan)
    b0_ts_T = np.empty(shape=(t_periods,)); b0_ts_T.fill(np.nan)
    c0_ts_T = np.empty(shape=(t_periods,)); c0_ts_T.fill(np.nan)
    BENs_M_T = np.empty(shape=(m_players, t_periods))                           # shape (M_PLAYERS, T_PERIODS)
    CSTs_M_T = np.empty(shape=(m_players, t_periods))
    CC_is_M = np.empty(shape=(m_players,)); CC_is_M.fill(np.nan)
    BB_is_M = np.empty(shape=(m_players,)); BB_is_M.fill(np.nan)
    RU_is_M = np.empty(shape=(m_players,)); RU_is_M.fill(np.nan)
    CC_is_M_T = np.empty(shape=(m_players, t_periods)); CC_is_M_T.fill(np.nan)
    BB_is_M_T = np.empty(shape=(m_players, t_periods)); BB_is_M_T.fill(np.nan)
    RU_is_M_T = np.empty(shape=(m_players, t_periods)); RU_is_M_T.fill(np.nan)
    pi_hp_minus_T = np.empty(shape=(t_periods,)); pi_hp_minus_T.fill(np.nan)
    pi_hp_plus_T = np.empty(shape=(t_periods,)); pi_hp_plus_T.fill(np.nan)
    
    arr_pl_M_T_vars_modif = arr_pl_M_T_vars_init.copy()
    arr_pl_M_T_vars_modif[:,:,fct_aux.AUTOMATE_INDEX_ATTRS["Si_minus"]] = np.nan
    arr_pl_M_T_vars_modif[:,:,fct_aux.AUTOMATE_INDEX_ATTRS["Si_plus"]] = np.nan
    arr_pl_M_T_vars_modif[:,:,fct_aux.AUTOMATE_INDEX_ATTRS["bg_i"]] = 0
    arr_pl_M_T_vars_modif[:,:,fct_aux.AUTOMATE_INDEX_ATTRS["u_i"]] = 0
    arr_pl_M_T_vars_modif[:,:,fct_aux.AUTOMATE_INDEX_ATTRS["S1_p_i_j_k"]] = 0.5
    arr_pl_M_T_vars_modif[:,:,fct_aux.AUTOMATE_INDEX_ATTRS["S2_p_i_j_k"]] = 0.5
    arr_pl_M_T_vars_modif[:,:,fct_aux.AUTOMATE_INDEX_ATTRS["non_playing_players"]] \
        = fct_aux.NON_PLAYING_PLAYERS["PLAY"]
        
    dico_id_players = {"players":[fct_aux.RACINE_PLAYER+"_"+str(num_pl_i) 
                                  for num_pl_i in range(0, m_players)]}
    df_nash = pd.DataFrame.from_dict(dico_id_players)
        
    # _________  creation des arrays pour chaque algo  _______________________
    arr_pl_M_T_vars_modif_BADBF = None
    arr_pl_M_T_vars_modif_BESTBF = None
    arr_pl_M_T_vars_modif_MIDBF = None
    
    pi_sg_plus_T_BESTBF = pi_sg_plus_T.copy()
    pi_sg_minus_T_BESTBF = pi_sg_minus_T.copy()
    pi_0_plus_T_BESTBF = pi_0_plus_T.copy()
    pi_0_minus_T_BESTBF = pi_0_minus_T.copy()
    B_is_M_BESTBF = B_is_M.copy()
    C_is_M_BESTBF = C_is_M.copy()
    b0_ts_T_BESTBF = b0_ts_T.copy()
    c0_ts_T_BESTBF = c0_ts_T.copy()
    BENs_M_T_BESTBF = BENs_M_T.copy()
    CSTs_M_T_BESTBF = CSTs_M_T.copy()
    CC_is_M_BESTBF = CC_is_M.copy()
    BB_is_M_BESTBF = BB_is_M.copy()
    RU_is_M_BESTBF = RU_is_M.copy()
    C_is_M_BESTBF = CC_is_M.copy()
    B_is_M_BESTBF = BB_is_M.copy()
    BB_is_M_T_BESTBF = BB_is_M_T.copy()
    CC_is_M_T_BESTBF = CC_is_M_T.copy()
    RU_is_M_T_BESTBF = RU_is_M_T.copy()
    B_is_M_T_BESTBF = B_is_M_T.copy()
    C_is_M_T_BESTBF = C_is_M_T.copy()
    dico_modes_profs_by_players_t_BESTBF = dict()
    df_nash_BESTBF = df_nash.copy()
    
    pi_sg_plus_T_BADBF = pi_sg_plus_T.copy()
    pi_sg_minus_T_BADBF = pi_sg_minus_T.copy()
    pi_0_plus_T_BADBF = pi_0_plus_T.copy()
    pi_0_minus_T_BADBF = pi_0_minus_T.copy()
    B_is_M_BADBF = B_is_M.copy()
    C_is_M_BADBF = C_is_M.copy()
    b0_ts_T_BADBF = b0_ts_T.copy()
    c0_ts_T_BADBF = c0_ts_T.copy()
    BENs_M_T_BADBF = BENs_M_T.copy()
    CSTs_M_T_BADBF = CSTs_M_T.copy()
    CC_is_M_BADBF = CC_is_M.copy()
    BB_is_M_BADBF = BB_is_M.copy()
    RU_is_M_BADBF = RU_is_M.copy()
    C_is_M_BADBF = CC_is_M.copy()
    B_is_M_BADBF = BB_is_M.copy()
    BB_is_M_T_BADBF = BB_is_M_T.copy()
    CC_is_M_T_BADBF = CC_is_M_T.copy()
    RU_is_M_T_BADBF = RU_is_M_T.copy()
    B_is_M_T_BADBF = B_is_M_T.copy()
    C_is_M_T_BADBF = C_is_M_T.copy()
    dico_modes_profs_by_players_t_BADBF = dict()
    df_nash_BADBF = df_nash.copy()
    
    pi_sg_plus_T_MIDBF = pi_sg_plus_T.copy()
    pi_sg_minus_T_MIDBF = pi_sg_minus_T.copy()
    pi_0_plus_T_MIDBF = pi_0_plus_T.copy()
    pi_0_minus_T_MIDBF = pi_0_minus_T.copy()
    B_is_M_MIDBF = B_is_M.copy()
    C_is_M_MIDBF = C_is_M.copy()
    b0_ts_T_MIDBF = b0_ts_T.copy()
    c0_ts_T_MIDBF = c0_ts_T.copy()
    BENs_M_T_MIDBF = BENs_M_T.copy()
    CSTs_M_T_MIDBF = CSTs_M_T.copy()
    CC_is_M_MIDBF = CC_is_M.copy()
    BB_is_M_MIDBF = BB_is_M.copy()
    RU_is_M_MIDBF = RU_is_M.copy()
    C_is_M_MIDBF = CC_is_M.copy()
    B_is_M_MIDBF = BB_is_M.copy()
    BB_is_M_T_MIDBF = BB_is_M_T.copy()
    CC_is_M_T_MIDBF = CC_is_M_T.copy()
    RU_is_M_T_MIDBF = RU_is_M_T.copy()
    B_is_M_T_MIDBF = B_is_M_T.copy()
    C_is_M_T_MIDBF = C_is_M_T.copy()
    dico_modes_profs_by_players_t_MIDBF = dict()
    df_nash_MIDBF = df_nash.copy()
    
    arr_pl_M_T_vars_modif_BADBF = None
    arr_pl_M_T_vars_modif_BESTBF = None
    arr_pl_M_T_vars_modif_MIDBF = None
    arr_pl_M_T_vars_modif_BADBF = arr_pl_M_T_vars_modif.copy()
    arr_pl_M_T_vars_modif_BESTBF = arr_pl_M_T_vars_modif.copy()
    arr_pl_M_T_vars_modif_MIDBF = arr_pl_M_T_vars_modif.copy()
        
    
    # ____      game beginning for all t_period ---> debut      _____
    pi_sg_plus_t0_minus_1 = pi_hp_plus-1
    pi_sg_minus_t0_minus_1 = pi_hp_minus-1
    pi_sg_plus_t_minus_1, pi_sg_minus_t_minus_1 = 0, 0
    pi_sg_plus_t, pi_sg_minus_t = None, None
        
    for t in range(0, t_periods):
        print("----- t = {} , free memory={}% ------ ".format(
            t, list(psutil.virtual_memory())[2]))
        pi_hp_plus_t, pi_hp_minus_t = None, None
        if manual_debug:
            pi_sg_plus_t = fct_aux.MANUEL_DBG_PI_SG_PLUS_T_K #8
            pi_sg_minus_t = fct_aux.MANUEL_DBG_PI_SG_MINUS_T_K #10
            pi_0_plus_t = fct_aux.MANUEL_DBG_PI_0_PLUS_T_K #2 
            pi_0_minus_t = fct_aux.MANUEL_DBG_PI_0_MINUS_T_K #3
        else:
            q_t_minus, q_t_plus = fct_aux.compute_upper_bound_quantity_energy(
                                    arr_pl_M_T_vars_modif, t)
            phi_hp_minus_t = fct_aux.compute_cost_energy_bought_by_SG_2_HP(
                                pi_hp_minus=pi_hp_minus, 
                                quantity=q_t_minus,
                                b=b)
            phi_hp_plus_t = fct_aux.compute_benefit_energy_sold_by_SG_2_HP(
                                pi_hp_plus=pi_hp_plus, 
                                quantity=q_t_plus,
                                a=a)
            pi_hp_minus_t = round(phi_hp_minus_t/q_t_minus, fct_aux.N_DECIMALS) \
                            if q_t_minus != 0 \
                            else 0
            pi_hp_plus_t = round(phi_hp_plus_t/q_t_plus, fct_aux.N_DECIMALS) \
                            if q_t_plus != 0 \
                            else 0
            if t == 0:
                pi_sg_plus_t0_minus_1 = pi_hp_plus_t - 1
                pi_sg_minus_t0_minus_1 = pi_hp_minus_t - 1
            pi_sg_plus_t_minus_1 = pi_sg_plus_t0_minus_1 if t == 0 \
                                                         else pi_sg_plus_t
            pi_sg_minus_t_minus_1 = pi_sg_minus_t0_minus_1 if t == 0 \
                                                            else pi_sg_minus_t
            
            print("q_t-={}, phi_hp-={}, pi_hp-={}, pi_sg-_t-1={}, ".format(q_t_minus, phi_hp_minus_t, pi_hp_minus_t, pi_sg_minus_t_minus_1))
            print("q_t+={}, phi_hp+={}, pi_hp+={}, pi_sg+_t-1={}".format(q_t_plus, phi_hp_plus_t, pi_hp_plus_t, pi_sg_plus_t_minus_1))
            
            pi_0_plus_t = None
            if pi_hp_minus_t > 0:
                pi_0_plus_t = round(pi_sg_minus_t_minus_1*pi_hp_plus_t/pi_hp_minus_t, 
                                    fct_aux.N_DECIMALS)
            else:
                pi_0_plus_t = 0
            pi_0_plus_t =  pi_0_plus_t if t > 0 else fct_aux.PI_0_PLUS_INIT #4
                                
            pi_0_minus_t = pi_sg_minus_t_minus_1 \
                            if t > 0 \
                            else fct_aux.PI_0_MINUS_INIT #3
            print("t={}, pi_0_plus_t={}, pi_0_minus_t={}".format(t, pi_0_plus_t, pi_0_minus_t))
               
        pi_0_plus_T[t] = pi_0_plus_t
        pi_0_minus_T[t] = pi_0_minus_t
        pi_hp_plus_T[t] = pi_hp_plus_t
        pi_hp_minus_T[t] = pi_hp_minus_t
        pi_sg_plus_T[t] = pi_sg_plus_t_minus_1
        pi_sg_minus_T[t] = pi_sg_minus_t_minus_1
        
               
        arr_pl_M_t_vars_init = arr_pl_M_T_vars_modif[:,t,:].copy()
        arr_pl_M_t_plus_1_vars_init = arr_pl_M_T_vars_modif[:,t+1,:].copy() \
                                        if t+1 < t_periods \
                                        else arr_pl_M_T_vars_modif[:,t,:].copy()
        arr_pl_M_t_minus_1_vars_init = arr_pl_M_T_vars_modif[:,t-1,:].copy() \
                                        if t-1 >= 0 \
                                        else arr_pl_M_T_vars_modif[:,t,:].copy()                        
        
        arr_pl_M_t_vars_modif = compute_gamma_state_4_period_t(
                                arr_pl_M_t_K_vars=arr_pl_M_t_vars_init,
                                arr_pl_M_t_minus_1_K_vars=arr_pl_M_t_minus_1_vars_init,
                                arr_pl_M_t_plus_1_K_vars=arr_pl_M_t_plus_1_vars_init,
                                t=t,
                                pi_0_plus=pi_0_plus_t, pi_0_minus=pi_0_minus_t,
                                pi_hp_plus_t=pi_hp_plus_t, pi_hp_minus_t=pi_hp_minus_t,
                                m_players=m_players,
                                t_periods=t_periods,
                                gamma_version=gamma_version,
                                manual_debug=manual_debug,
                                dbg=dbg)
        arr_pl_M_T_vars_modif_BADBF[:,t,:] = arr_pl_M_t_vars_modif.copy()
        arr_pl_M_T_vars_modif_BESTBF[:,t,:] = arr_pl_M_t_vars_modif.copy()
        arr_pl_M_T_vars_modif_MIDBF[:,t,:] = arr_pl_M_t_vars_modif.copy()
        
        # quantity of energies (prod_is, cons_is) from 0 to t-2 to get values 
        # for t-1 periods
        sum_diff_pos_minus_0_t_minus_2 = None                                  # sum of the positive difference btw cons_is and prod_is from 0 to t-2 
        sum_diff_pos_plus_0_t_minus_2 = None                                   # sum of the positive difference btw prod_is and cons_is from 0 to t-2
        sum_cons_is_0_t_minus_2 = None                                         # sum of the cons of all players from 0 to t-2
        sum_prod_is_0_t_minus_2 = None                                         # sum of the prod of all players from 0 to t-2
        
        sum_diff_pos_minus_0_t_minus_2, \
        sum_diff_pos_plus_0_t_minus_2, \
        sum_cons_is_0_t_minus_2, \
        sum_prod_is_0_t_minus_2 \
            = get_sum_cons_prod_from_0_t_minus_2(arr_pl_M_T_vars_modif,t)
        print("t={}, sum_diff_pos_minus_0_t_minus_2={}, sum_diff_pos_plus_0_t_minus_2={}, sum_cons_is_0_t_minus_2={}, sum_prod_is_0_t_minus_2={}".format(t,sum_diff_pos_minus_0_t_minus_2,
                        sum_diff_pos_plus_0_t_minus_2,
                        sum_cons_is_0_t_minus_2, 
                        sum_prod_is_0_t_minus_2))
        
            
        # balanced player game at instant t    
        list_dico_modes_profs_by_players_t_best = list()
        list_dico_modes_profs_by_players_t_bad = list()
        list_dico_modes_profs_by_players_t_mid = list()
        
        list_dico_modes_profs_by_players_t_best, \
        list_dico_modes_profs_by_players_t_bad, \
        list_dico_modes_profs_by_players_t_mid\
            = generer_balanced_players_4_modes_profils(
                arr_pl_M_t_vars_modif, 
                m_players, t,
                sum_diff_pos_minus_0_t_minus_2,
                sum_diff_pos_plus_0_t_minus_2,
                sum_cons_is_0_t_minus_2,                             
                sum_prod_is_0_t_minus_2,
                pi_hp_plus, pi_hp_minus,
                a, b,
                pi_0_plus_t, pi_0_minus_t,
                manual_debug, dbg)
        
        # appliquer sur chaque algo BEST, BAD, MIDDLE
        for algo_name in fct_aux.ALGO_NAMES_BF:
            list_dico_modes_profs_by_players_t_algo = dict()
            if algo_name == fct_aux.ALGO_NAMES_BF[0]:                          # BEST-BRUTE-FORCE
                list_dico_modes_profs_by_players_t_algo \
                    = list_dico_modes_profs_by_players_t_best
                arr_pl_M_t_vars_modif_algo \
                    = arr_pl_M_T_vars_modif_BESTBF[:,t,:].copy()
                df_nash_algo = df_nash_BESTBF.copy()
                
            elif algo_name == fct_aux.ALGO_NAMES_BF[1]:                        # BAD-BRUTE-FORCE
                list_dico_modes_profs_by_players_t_algo \
                    = list_dico_modes_profs_by_players_t_bad
                arr_pl_M_t_vars_modif_algo \
                    = arr_pl_M_T_vars_modif_BADBF[:,t,:].copy()
                df_nash_algo = df_nash_BADBF.copy()
                
            elif algo_name == fct_aux.ALGO_NAMES_BF[2]:                        # MIDDLE-BRUTE-FORCE
                list_dico_modes_profs_by_players_t_algo \
                    = list_dico_modes_profs_by_players_t_mid
                arr_pl_M_t_vars_modif_algo \
                    = arr_pl_M_T_vars_modif_MIDBF[:,t,:].copy()
                df_nash_algo = df_nash_MIDBF.copy()
            
            rd_key = None
            if len(list_dico_modes_profs_by_players_t_algo) == 1:
                rd_key = 0
            else:
                rd_key = np.random.randint(
                            0, 
                            len(list_dico_modes_profs_by_players_t_algo))
            
            id_cpt_xxx, dico_mode_prof_by_players_algo \
                = list_dico_modes_profs_by_players_t_algo[rd_key] 
            print("{}, rd_key={}, cpt_xxx={}".format(algo_name, rd_key, id_cpt_xxx))
            
            bens_t_algo = dico_mode_prof_by_players_algo["bens_t"]
            csts_t_algo = dico_mode_prof_by_players_algo["csts_t"]
            Perf_t_algo = dico_mode_prof_by_players_algo["Perf_t"]
            b0_t_algo = dico_mode_prof_by_players_algo["b0_t"]
            c0_t_algo = dico_mode_prof_by_players_algo["c0_t"]
            Out_sg_algo = dico_mode_prof_by_players_algo["Out_sg"]
            In_sg_algo = dico_mode_prof_by_players_algo["In_sg"]
            pi_sg_plus_t_algo = dico_mode_prof_by_players_algo["pi_sg_plus_t"]
            pi_sg_minus_t_algo = dico_mode_prof_by_players_algo["pi_sg_minus_t"]
            pi_0_plus_t_algo = dico_mode_prof_by_players_algo["pi_0_plus_t"]
            pi_0_minus_t_algo = dico_mode_prof_by_players_algo["pi_0_minus_t"]
            mode_profile = dico_mode_prof_by_players_algo["mode_profile"]
            
            pi_sg_plus_t = pi_sg_plus_t_algo
            pi_sg_minus_t = pi_sg_minus_t_algo
            
            arr_pl_M_t_vars_modif_algo[:, 
                                  fct_aux.AUTOMATE_INDEX_ATTRS["mode_i"]] \
                = mode_profile
            
            print("mode_profile={}, mode_is={}".format(mode_profile, 
                    arr_pl_M_t_vars_modif_algo[:,fct_aux.AUTOMATE_INDEX_ATTRS["mode_i"]]))
            print("state_is={} ".format( 
                    arr_pl_M_t_vars_modif_algo[:,fct_aux.AUTOMATE_INDEX_ATTRS["state_i"]]))
            
            arr_pl_M_t_vars_modif_algo = balanced_player_game_4_mode_profil(
                                             arr_pl_M_t_vars_modif_algo, 
                                             m_players,
                                             dbg)
            ## test if there are the same values like these in dico_mode_prof_by_players
            In_sg_new, Out_sg_new, b0_t_new, c0_t_new = None, None, None, None
            bens_t_new, csts_t_new = None, None
            pi_sg_plus_t_new, pi_sg_minus_t_new = None, None
            In_sg_new, Out_sg_new, \
            b0_t_new, c0_t_new, \
            bens_t_new, csts_t_new, \
            pi_sg_plus_t_new, pi_sg_minus_t_new \
                = compute_prices_inside_SG(
                    arr_pl_M_t_vars_modif_algo, 
                    sum_diff_pos_minus_0_t_minus_2,
                    sum_diff_pos_plus_0_t_minus_2,
                    sum_cons_is_0_t_minus_2,                             
                    sum_prod_is_0_t_minus_2,
                    pi_hp_plus, pi_hp_minus,
                    a, b,
                    pi_0_plus_t, pi_0_minus_t,
                    manual_debug, dbg)
            bens_csts_t_new = bens_t_new - csts_t_new
            Perf_t_new = np.sum(bens_csts_t_new, axis=0)
            ##### verification of best key quality 
            diff = np.abs(Perf_t_new - Perf_t_algo)
            print(" Perf_t_algo == Perf_t_new --> OK (diff={}) ".format(diff)) \
                if diff < 0.1 \
                else print("Perf_t_algo != Perf_t_new --> NOK (diff={}) \n"\
                           .format(diff))     
            print("b0_t={}, c0_t={}, Out_sg={},In_sg={}, pi_hp_minus_t={}, pi_hp_plus_t={}\n".format(
                    b0_t_algo, c0_t_algo, Out_sg_algo, In_sg_algo, 
                    pi_hp_minus_t, pi_hp_plus_t))
            
            # pi_sg_{plus,minus} of shape (T_PERIODS,)
            if np.isnan(pi_sg_plus_t_algo):
                pi_sg_plus_t_algo = 0
            if np.isnan(pi_sg_minus_t_algo):
                pi_sg_minus_t_algo = 0
                
            
            # checkout NASH equilibrium
            bens_csts_M_t = bens_t_algo - csts_t_algo
            df_nash_t = None
            df_nash_t = checkout_nash_4_profils_by_periods(
                            arr_pl_M_t_vars_modif_algo.copy(),
                            arr_pl_M_T_vars_init[:,t,:],
                            sum_diff_pos_minus_0_t_minus_2,
                            sum_diff_pos_plus_0_t_minus_2,
                            sum_cons_is_0_t_minus_2,                             
                            sum_prod_is_0_t_minus_2,
                            pi_hp_plus=pi_hp_plus, pi_hp_minus=pi_hp_minus, 
                            a=a, b=b,
                            pi_0_minus_t=pi_0_minus_t, pi_0_plus_t=pi_0_plus_t, 
                            bens_csts_M_t=bens_csts_M_t,
                            m_players=m_players,
                            t=t,
                            manual_debug=manual_debug,
                            dbg=dbg)
            df_nash_algo = pd.merge(df_nash_algo, df_nash_t, on='players', 
                                    how='outer')
            
            #_______     save arr_M_t_vars at t in dataframe : debut    _______
            df_arr_M_t_vars_modif_algo \
                = pd.DataFrame(arr_pl_M_t_vars_modif_algo, 
                                columns=fct_aux.AUTOMATE_INDEX_ATTRS.keys(),
                                index=dico_id_players["players"])
            path_to_save_M_t_vars_modif_algo = path_to_save
            msg = "pi_hp_plus_"+str(pi_hp_plus)+"_pi_hp_minus_"+str(pi_hp_minus)
            if "simu_DDMM_HHMM" in path_to_save:
                path_to_save_M_t_vars_modif_algo \
                    = os.path.join(name_dir, "simu_"+date_hhmm,
                                   msg, algo_name, "intermediate_t"
                                   )
            Path(path_to_save_M_t_vars_modif_algo).mkdir(parents=True, 
                                                         exist_ok=True)
                
            path_2_xls_df_arr_M_t_vars_modif_algo \
                = os.path.join(
                    path_to_save_M_t_vars_modif_algo,
                      "arr_M_T_vars_{}.xlsx".format(algo_name)
                    )
            if not os.path.isfile(path_2_xls_df_arr_M_t_vars_modif_algo):
                df_arr_M_t_vars_modif_algo.to_excel(
                    path_2_xls_df_arr_M_t_vars_modif_algo,
                    sheet_name="t{}".format(t),
                    index=True)
            else:
                book = load_workbook(filename=path_2_xls_df_arr_M_t_vars_modif_algo)
                with pd.ExcelWriter(path_2_xls_df_arr_M_t_vars_modif_algo, 
                                    engine='openpyxl') as writer:
                    writer.book = book
                    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)    
                
                    ## Your dataframe to append. 
                    df_arr_M_t_vars_modif_algo.to_excel(writer, "t{}".format(t))  
                
                    writer.save() 
            #_______     save arr_M_t_vars at t in dataframe : fin     _______
            
            #___________ update saving variables : debut ______________________
            if algo_name == fct_aux.ALGO_NAMES_BF[0]:                          # BEST-BRUTE-FORCE
                arr_pl_M_T_vars_modif_BESTBF[:,t,:] \
                    = arr_pl_M_t_vars_modif_algo 
                b0_ts_T_BESTBF, c0_ts_T_BESTBF, \
                BENs_M_T_BESTBF, CSTs_M_T_BESTBF, \
                pi_sg_plus_T_BESTBF, pi_sg_minus_T_BESTBF, \
                pi_0_plus_T_BESTBF, pi_0_minus_T_BESTBF, \
                df_nash_BESTBF \
                    = update_saving_variables(t, 
                        b0_ts_T_BESTBF, b0_t_algo,
                        c0_ts_T_BESTBF, c0_t_algo,
                        BENs_M_T_BESTBF, bens_t_algo,
                        CSTs_M_T_BESTBF, csts_t_algo,
                        pi_sg_plus_T_BESTBF, pi_sg_plus_t_algo,
                        pi_sg_minus_T_BESTBF, pi_sg_minus_t_algo,
                        pi_0_plus_T_BESTBF, pi_0_plus_t_algo,
                        pi_0_minus_T_BESTBF, pi_0_minus_t_algo,
                        df_nash_BESTBF, df_nash_algo
                        )
                BB_is_M_T_BESTBF, CC_is_M_T_BESTBF, RU_is_M_T_BESTBF, \
                B_is_M_T_BESTBF, C_is_M_T_BESTBF \
                    = compute_prices_variables(
                        arr_pl_M_T_vars_modif_BESTBF, t,
                        b0_ts_T_BESTBF, c0_ts_T_BESTBF, 
                        pi_sg_plus_T_BESTBF, pi_sg_minus_T_BESTBF,
                        pi_0_plus_T_BESTBF, pi_0_minus_T_BESTBF
                        )
                dico_modes_profs_by_players_t_BESTBF[t] \
                    = dico_mode_prof_by_players_algo
    
            elif algo_name == fct_aux.ALGO_NAMES_BF[1]:                        # BAD-BRUTE-FORCE
                arr_pl_M_T_vars_modif_BADBF[:,t,:] \
                    = arr_pl_M_t_vars_modif_algo 
                b0_ts_T_BADBF, c0_ts_T_BADBF, \
                BENs_M_T_BADBF, CSTs_M_T_BADBF, \
                pi_sg_plus_T_BADBF, pi_sg_minus_T_BADBF, \
                pi_0_plus_T_BADBF, pi_0_minus_T_BADBF, \
                df_nash_BADBF \
                    = update_saving_variables(t, 
                        b0_ts_T_BADBF, b0_t_algo,
                        c0_ts_T_BADBF, c0_t_algo,
                        BENs_M_T_BADBF, bens_t_algo,
                        CSTs_M_T_BADBF, csts_t_algo,
                        pi_sg_plus_T_BADBF, pi_sg_plus_t_algo,
                        pi_sg_minus_T_BADBF, pi_sg_minus_t_algo,
                        pi_0_plus_T_BADBF, pi_0_plus_t_algo,
                        pi_0_minus_T_BADBF, pi_0_minus_t_algo,
                        df_nash_BADBF, df_nash_algo
                        )
                BB_is_M_T_BADBF, CC_is_M_T_BADBF, RU_is_M_T_BADBF, \
                B_is_M_T_BADBF, C_is_M_T_BADBF \
                    = compute_prices_variables(
                        arr_pl_M_T_vars_modif_BADBF, t,
                        b0_ts_T_BADBF, c0_ts_T_BADBF,
                        pi_sg_plus_T_BADBF, pi_sg_minus_T_BADBF,
                        pi_0_plus_T_BADBF, pi_0_minus_T_BADBF
                        )
                dico_modes_profs_by_players_t_BADBF[t] \
                    = dico_mode_prof_by_players_algo    
                    
            elif algo_name == fct_aux.ALGO_NAMES_BF[2]:                        # MIDDLE-BRUTE-FORCE
                arr_pl_M_T_vars_modif_MIDBF[:,t,:] \
                    = arr_pl_M_t_vars_modif_algo 
                b0_ts_T_MIDBF, c0_ts_T_MIDBF, \
                BENs_M_T_MIDBF, CSTs_M_T_MIDBF, \
                pi_sg_plus_T_MIDBF, pi_sg_minus_T_MIDBF, \
                pi_0_plus_T_MIDBF, pi_0_minus_T_MIDBF, \
                df_nash_MIDBF \
                    = update_saving_variables(t,
                        b0_ts_T_MIDBF, b0_t_algo,
                        c0_ts_T_MIDBF, c0_t_algo,
                        BENs_M_T_MIDBF, bens_t_new,
                        CSTs_M_T_MIDBF, csts_t_new,
                        pi_sg_plus_T_MIDBF, pi_sg_plus_t_algo,
                        pi_sg_minus_T_MIDBF, pi_sg_minus_t_algo,
                        pi_0_plus_T_MIDBF, pi_0_plus_t_algo,
                        pi_0_minus_T_MIDBF, pi_0_minus_t_algo,
                        df_nash_MIDBF, df_nash_algo
                        )
                BB_is_M_T_MIDBF, CC_is_M_T_MIDBF, RU_is_M_T_MIDBF, \
                B_is_M_T_MIDBF, C_is_M_T_MIDBF \
                    = compute_prices_variables(
                        arr_pl_M_T_vars_modif_MIDBF, t,
                        b0_ts_T_MIDBF, c0_ts_T_MIDBF,
                        pi_sg_plus_T_MIDBF, pi_sg_minus_T_MIDBF,
                        pi_0_plus_T_MIDBF, pi_0_minus_T_MIDBF, 
                        )
                dico_modes_profs_by_players_t_MIDBF[t] \
                    = dico_mode_prof_by_players_algo
            #___________ update saving variables : fin   ______________________
        
        print("----- t={} After running free memory={}% ------ ".format(
            t, list(psutil.virtual_memory())[2]))
        
    # __________        compute real prices variables         _________________
    
    B_is_M_BESTBF = np.sum(B_is_M_T_BESTBF, axis=1)
    C_is_M_BESTBF = np.sum(C_is_M_T_BESTBF, axis=1)
    BB_is_M_BESTBF = np.sum(BB_is_M_T_BESTBF, axis=1) 
    CC_is_M_BESTBF = np.sum(CC_is_M_T_BESTBF, axis=1) 
    RU_is_M_BESTBF = np.sum(RU_is_M_T_BESTBF, axis=1)

    B_is_M_BADBF = np.sum(B_is_M_T_BADBF, axis=1)
    C_is_M_BADBF = np.sum(C_is_M_T_BADBF, axis=1)
    BB_is_M_BADBF = np.sum(BB_is_M_T_BADBF, axis=1) 
    CC_is_M_BADBF = np.sum(CC_is_M_T_BADBF, axis=1) 
    RU_is_M_BADBF = np.sum(RU_is_M_T_BADBF, axis=1) 

    B_is_M_MIDBF = np.sum(B_is_M_T_MIDBF, axis=1)
    C_is_M_MIDBF = np.sum(C_is_M_T_MIDBF, axis=1)
    BB_is_M_MIDBF = np.sum(BB_is_M_T_MIDBF, axis=1) 
    CC_is_M_MIDBF = np.sum(CC_is_M_T_MIDBF, axis=1) 
    RU_is_M_MIDBF = np.sum(RU_is_M_T_MIDBF, axis=1)   

    # __________        compute real prices variables         _________________ 
       

    #_______      save computed variables locally from algo_name     __________
    msg = "pi_hp_plus_"+str(pi_hp_plus)+"_pi_hp_minus_"+str(pi_hp_minus)
    
    print("path_to_save={}".format(path_to_save))
    algo_name = fct_aux.ALGO_NAMES_BF[0]
    if "simu_DDMM_HHMM" in path_to_save:
        path_to_save = os.path.join(name_dir, "simu_"+date_hhmm,
                                    msg, algo_name
                                    )
    Path(path_to_save).mkdir(parents=True, exist_ok=True)
    df_nash_BESTBF.to_excel(os.path.join(
                *[path_to_save,
                  "resume_verify_Nash_equilibrium_{}.xlsx".format(algo_name)]), 
                index=False)

    fct_aux.save_variables(path_to_save, arr_pl_M_T_vars_modif_BESTBF, 
               b0_ts_T_BESTBF, c0_ts_T_BESTBF, B_is_M_BESTBF, C_is_M_BESTBF, 
               BENs_M_T_BESTBF, CSTs_M_T_BESTBF, 
               BB_is_M_BESTBF, CC_is_M_BESTBF, RU_is_M_BESTBF, 
               pi_sg_minus_T_BESTBF, pi_sg_plus_T_BESTBF, 
               pi_0_minus_T_BESTBF, pi_0_plus_T_BESTBF,
               pi_hp_plus_T, pi_hp_minus_T,
               dico_modes_profs_by_players_t_BESTBF, 
               algo=algo_name,
               dico_best_steps=dict())
    turn_dico_stats_res_into_df_BF(
          dico_modes_profs_players_algo= dico_modes_profs_by_players_t_BESTBF, 
          path_to_save = path_to_save, 
          t_periods = t_periods, 
          manual_debug = manual_debug, 
          algo_name = algo_name)
    
    algo_name = fct_aux.ALGO_NAMES_BF[1]
    if "simu_DDMM_HHMM" in path_to_save:
        path_to_save = os.path.join(name_dir, "simu_"+date_hhmm,
                                    msg, algo_name
                                    )
    Path(path_to_save).mkdir(parents=True, exist_ok=True)
    df_nash_BADBF.to_excel(os.path.join(
                *[path_to_save,
                  "resume_verify_Nash_equilibrium_{}.xlsx".format(algo_name)]), 
                index=False)
    fct_aux.save_variables(path_to_save, arr_pl_M_T_vars_modif_BADBF, 
                   b0_ts_T_BADBF, c0_ts_T_BADBF, B_is_M_BADBF, C_is_M_BADBF, 
                   BENs_M_T_BADBF, CSTs_M_T_BADBF, 
                   BB_is_M_BADBF, CC_is_M_BADBF, RU_is_M_BADBF, 
                   pi_sg_minus_T_BADBF, pi_sg_plus_T_BADBF, 
                   pi_0_minus_T_BADBF, pi_0_plus_T_BADBF,
                   pi_hp_plus_T, pi_hp_minus_T,
                   dico_modes_profs_by_players_t_BADBF, 
                   algo=algo_name,
                   dico_best_steps=dict())
    turn_dico_stats_res_into_df_BF(
          dico_modes_profs_players_algo= dico_modes_profs_by_players_t_BADBF, 
          path_to_save = path_to_save, 
          t_periods = t_periods, 
          manual_debug = manual_debug, 
          algo_name = algo_name)
    
    algo_name = fct_aux.ALGO_NAMES_BF[2]
    if "simu_DDMM_HHMM" in path_to_save:
        path_to_save = os.path.join(name_dir, "simu_"+date_hhmm,
                                    msg, algo_name
                                    )
    Path(path_to_save).mkdir(parents=True, exist_ok=True)
    df_nash_MIDBF.to_excel(os.path.join(
                *[path_to_save,
                  "resume_verify_Nash_equilibrium_{}.xlsx".format(algo_name)]), 
                index=False)
    fct_aux.save_variables(path_to_save, arr_pl_M_T_vars_modif_MIDBF, 
                   b0_ts_T_MIDBF, c0_ts_T_MIDBF, B_is_M_MIDBF, C_is_M_MIDBF, 
                   BENs_M_T_MIDBF, CSTs_M_T_MIDBF, 
                   BB_is_M_MIDBF, CC_is_M_MIDBF, RU_is_M_MIDBF, 
                   pi_sg_minus_T_MIDBF, pi_sg_plus_T_MIDBF, 
                   pi_0_minus_T_MIDBF, pi_0_plus_T_MIDBF,
                   pi_hp_plus_T, pi_hp_minus_T,
                   dico_modes_profs_by_players_t_MIDBF, 
                   algo=algo_name,
                   dico_best_steps=dict())
    turn_dico_stats_res_into_df_BF(
          dico_modes_profs_players_algo= dico_modes_profs_by_players_t_MIDBF, 
          path_to_save = path_to_save, 
          t_periods = t_periods, 
          manual_debug = manual_debug, 
          algo_name = algo_name)
    return arr_pl_M_T_vars_modif
    
# __________       main function of Brute force   ---> fin        ____________

# __________       main function of One BF Algo   ---> debut         __________  
def bf_balanced_player_game_ONE_ALGO(arr_pl_M_T_vars_init, algo_name,
                                    pi_hp_plus=0.02, 
                                    pi_hp_minus=0.33,
                                    a=1, b=1,
                                    gamma_version=1,
                                    path_to_save="tests", 
                                    name_dir="tests", 
                                    date_hhmm="DDMM_HHMM",
                                    manual_debug=False, 
                                    criteria_bf="Perf_t", dbg=False):
    """
    """
    print("\n \n game: pi_hp_plus={}, pi_hp_minus={} ---> debut \n"\
          .format(pi_hp_plus, pi_hp_minus))
        
    m_players = arr_pl_M_T_vars_init.shape[0]
    t_periods = arr_pl_M_T_vars_init.shape[1]
    
    # _______ variables' initialization --> debut ________________
    pi_sg_plus_t, pi_sg_minus_t = 0, 0
    pi_sg_plus_T = np.empty(shape=(t_periods,))                                 # shape (T_PERIODS,)
    pi_sg_plus_T.fill(np.nan)
    pi_sg_minus_T = np.empty(shape=(t_periods,))                                # shape (T_PERIODS,)
    pi_sg_plus_T.fill(np.nan)
    pi_0_plus_t, pi_0_minus_t = 0, 0
    pi_0_plus_T = np.empty(shape=(t_periods,))                                  # shape (T_PERIODS,)
    pi_0_plus_T.fill(np.nan)
    pi_0_minus_T = np.empty(shape=(t_periods,))                                 # shape (T_PERIODS,)
    pi_0_minus_T.fill(np.nan)
    B_is_M = np.empty(shape=(m_players,))                                       # shape (M_PLAYERS, )
    B_is_M.fill(np.nan)
    C_is_M = np.empty(shape=(m_players,))                                       # shape (M_PLAYERS, )
    C_is_M.fill(np.nan)
    B_is_M_T = np.empty(shape=(m_players, t_periods))                           # shape (M_PLAYERS, )
    B_is_M_T.fill(np.nan)
    C_is_M_T = np.empty(shape=(m_players, t_periods))                           # shape (M_PLAYERS, )
    C_is_M_T.fill(np.nan)
    b0_ts_T = np.empty(shape=(t_periods,))                                      # shape (T_PERIODS,)
    b0_ts_T.fill(np.nan)
    c0_ts_T = np.empty(shape=(t_periods,))
    c0_ts_T.fill(np.nan)
    BENs_M_T = np.empty(shape=(m_players, t_periods))                           # shape (M_PLAYERS, T_PERIODS)
    CSTs_M_T = np.empty(shape=(m_players, t_periods))
    CC_is_M = np.empty(shape=(m_players,))                                      # shape (M_PLAYERS, )
    CC_is_M.fill(np.nan)
    BB_is_M = np.empty(shape=(m_players,))                                      # shape (M_PLAYERS, )
    BB_is_M.fill(np.nan)
    EB_is_M = np.empty(shape=(m_players,))                                      # shape (M_PLAYERS, )
    EB_is_M.fill(np.nan)
    CC_is_M_T = np.empty(shape=(m_players, t_periods))                          # shape (M_PLAYERS, )
    CC_is_M_T.fill(np.nan)
    BB_is_M_T = np.empty(shape=(m_players, t_periods))                          # shape (M_PLAYERS, )
    BB_is_M_T.fill(np.nan)
    EB_is_M_T = np.empty(shape=(m_players, t_periods))                          # shape (M_PLAYERS, )
    EB_is_M_T.fill(np.nan)
    pi_hp_minus_T = np.empty(shape=(t_periods,))
    pi_hp_plus_T = np.empty(shape=(t_periods,))
    
    arr_pl_M_T_vars_modif = arr_pl_M_T_vars_init.copy()
    arr_pl_M_T_vars_modif[:,:,fct_aux.AUTOMATE_INDEX_ATTRS["Si_minus"]] = np.nan
    arr_pl_M_T_vars_modif[:,:,fct_aux.AUTOMATE_INDEX_ATTRS["Si_plus"]] = np.nan
    arr_pl_M_T_vars_modif[:,:,fct_aux.AUTOMATE_INDEX_ATTRS["bg_i"]] = 0
    arr_pl_M_T_vars_modif[:,:,fct_aux.AUTOMATE_INDEX_ATTRS["u_i"]] = 0
    arr_pl_M_T_vars_modif[:,:,fct_aux.AUTOMATE_INDEX_ATTRS["S1_p_i_j_k"]] = 0.5
    arr_pl_M_T_vars_modif[:,:,fct_aux.AUTOMATE_INDEX_ATTRS["S2_p_i_j_k"]] = 0.5
    arr_pl_M_T_vars_modif[:,:,fct_aux.AUTOMATE_INDEX_ATTRS["non_playing_players"]] \
        = fct_aux.NON_PLAYING_PLAYERS["PLAY"]
        
    dico_id_players = {"players":[fct_aux.RACINE_PLAYER+"_"+str(num_pl_i) 
                                  for num_pl_i in range(0, m_players)]}
    df_nash = pd.DataFrame.from_dict(dico_id_players)
    
    # ____      game beginning for all t_period ---> debut      _____
   
    pi_sg_plus_t0_minus_1, pi_sg_minus_t0_minus_1 = None, None
    pi_sg_plus_t_minus_1, pi_sg_minus_t_minus_1 = None, None
    pi_sg_plus_t, pi_sg_minus_t = None, None
    pi_hp_plus_t, pi_hp_minus_t = None, None 
    
    dico_profils_bf, dico_best_profils_bf, dico_bad_profils_bf = dict(), dict(), dict()
    
    dico_modes_profs_by_players_t = dict() 
    for t in range(0, t_periods):
        print("----- t = {} , free memory={}% ------ ".format(
            t, list(psutil.virtual_memory())[2]))
        pi_hp_plus_t, pi_hp_minus_t = None, None
        pi_0_plus_t, pi_0_minus_t = None, None
        if manual_debug:
            pi_sg_plus_t = fct_aux.MANUEL_DBG_PI_SG_PLUS_T_K #8
            pi_sg_minus_t = fct_aux.MANUEL_DBG_PI_SG_MINUS_T_K #10
            pi_0_plus_t = fct_aux.MANUEL_DBG_PI_0_PLUS_T_K #2 
            pi_0_minus_t = fct_aux.MANUEL_DBG_PI_0_MINUS_T_K #3
            pi_hp_plus_t, pi_hp_minus_t = pi_hp_plus, pi_hp_minus
        else:
            q_t_minus, q_t_plus = fct_aux.compute_upper_bound_quantity_energy(
                                    arr_pl_M_T_vars_modif, t)
            phi_hp_minus_t = fct_aux.compute_cost_energy_bought_by_SG_2_HP(
                                pi_hp_minus=pi_hp_minus, 
                                quantity=q_t_minus,
                                b=b)
            phi_hp_plus_t = fct_aux.compute_benefit_energy_sold_by_SG_2_HP(
                                pi_hp_plus=pi_hp_plus, 
                                quantity=q_t_plus,
                                a=a)
            pi_hp_minus_t = round(phi_hp_minus_t/q_t_minus, fct_aux.N_DECIMALS) \
                            if q_t_minus != 0 \
                            else 0
            pi_hp_plus_t = round(phi_hp_plus_t/q_t_plus, fct_aux.N_DECIMALS) \
                            if q_t_plus != 0 \
                            else 0
            if t == 0:
                pi_sg_plus_t0_minus_1 = pi_hp_plus_t - 1
                pi_sg_minus_t0_minus_1 = pi_hp_minus_t - 1
            pi_sg_plus_t_minus_1 = pi_sg_plus_t0_minus_1 if t == 0 \
                                                         else pi_sg_plus_t
            pi_sg_minus_t_minus_1 = pi_sg_minus_t0_minus_1 if t == 0 \
                                                            else pi_sg_minus_t
            
            print("q_t-={}, phi_hp-={}, pi_hp-={}, pi_sg-_t-1={}, ".format(q_t_minus, phi_hp_minus_t, pi_hp_minus_t, pi_sg_minus_t_minus_1))
            print("q_t+={}, phi_hp+={}, pi_hp+={}, pi_sg+_t-1={}".format(q_t_plus, phi_hp_plus_t, pi_hp_plus_t, pi_sg_plus_t_minus_1))
            
            pi_0_plus_t = round(pi_sg_minus_t_minus_1*pi_hp_plus_t/pi_hp_minus_t, 
                                fct_aux.N_DECIMALS) \
                            if t > 0 \
                            else fct_aux.PI_0_PLUS_INIT #4
                                
            pi_0_minus_t = pi_sg_minus_t_minus_1 \
                            if t > 0 \
                            else fct_aux.PI_0_MINUS_INIT #3
            print("t={}, pi_0_plus_t={}, pi_0_minus_t={}".format(t, pi_0_plus_t, pi_0_minus_t))
                  
        pi_0_plus_T[t] = pi_0_plus_t
        pi_0_minus_T[t] = pi_0_minus_t
        pi_hp_plus_T[t] = pi_hp_plus_t
        pi_hp_minus_T[t] = pi_hp_minus_t
        pi_sg_plus_T[t] = pi_sg_plus_t_minus_1
        pi_sg_minus_T[t] = pi_sg_minus_t_minus_1
               
        arr_pl_M_t_vars_init = arr_pl_M_T_vars_modif[:,t,:].copy()
        arr_pl_M_t_plus_1_vars_init = arr_pl_M_T_vars_modif[:,t+1,:].copy() \
                                        if t+1 < t_periods \
                                        else arr_pl_M_T_vars_modif[:,t,:].copy()
        arr_pl_M_t_minus_1_vars_init = arr_pl_M_T_vars_modif[:,t-1,:].copy() \
                                        if t-1 >= 0 \
                                        else arr_pl_M_T_vars_modif[:,t,:].copy()                        
        
        print("Sis_init = {}".format(arr_pl_M_T_vars_modif[:,t,fct_aux.AUTOMATE_INDEX_ATTRS["Si"]]))
        print("Sis_+1 = {}".format(arr_pl_M_t_plus_1_vars_init[:,fct_aux.AUTOMATE_INDEX_ATTRS["Si"]]))
        print("Sis_-1 = {}".format(arr_pl_M_t_minus_1_vars_init[:,fct_aux.AUTOMATE_INDEX_ATTRS["Si"]]))
        
        arr_pl_M_t_vars_modif = compute_gamma_state_4_period_t(
                                arr_pl_M_t_K_vars=arr_pl_M_t_vars_init,
                                arr_pl_M_t_minus_1_K_vars=arr_pl_M_t_minus_1_vars_init,
                                arr_pl_M_t_plus_1_K_vars=arr_pl_M_t_plus_1_vars_init,
                                t=t,
                                pi_0_plus=pi_0_plus_t, pi_0_minus=pi_0_minus_t,
                                pi_hp_plus_t=pi_hp_plus_t, pi_hp_minus_t=pi_hp_minus_t,
                                m_players=m_players,
                                t_periods=t_periods,
                                gamma_version=gamma_version,
                                manual_debug=manual_debug,
                                dbg=dbg)
        
        print("Sis = {} \n".format(arr_pl_M_t_vars_modif[:,fct_aux.AUTOMATE_INDEX_ATTRS["Si"]]))

        
        # quantity of energies (prod_is, cons_is) from 0 to t-2 to get values 
        # for t-1 periods
        sum_diff_pos_minus_0_t_minus_2 = None                                  # sum of the positive difference btw cons_is and prod_is from 0 to t-2 
        sum_diff_pos_plus_0_t_minus_2 = None                                   # sum of the positive difference btw prod_is and cons_is from 0 to t-2
        sum_cons_is_0_t_minus_2 = None                                         # sum of the cons of all players from 0 to t-2
        sum_prod_is_0_t_minus_2 = None                                         # sum of the prod of all players from 0 to t-2
        
        sum_diff_pos_minus_0_t_minus_2, \
        sum_diff_pos_plus_0_t_minus_2, \
        sum_cons_is_0_t_minus_2, \
        sum_prod_is_0_t_minus_2 \
            = get_sum_cons_prod_from_0_t_minus_2(arr_pl_M_T_vars_modif,t)
        print("t={}, sum_diff_pos_minus_0_t_minus_2={}, sum_diff_pos_plus_0_t_minus_2={}, sum_cons_is_0_t_minus_2={}, sum_prod_is_0_t_minus_2={}".format(t,sum_diff_pos_minus_0_t_minus_2,
                        sum_diff_pos_plus_0_t_minus_2,
                        sum_cons_is_0_t_minus_2, 
                        sum_prod_is_0_t_minus_2))
        
            
        # balanced player game at instant t    
        list_dico_modes_profs_by_players_t_best = list()
        list_dico_modes_profs_by_players_t_bad = list()
        list_dico_modes_profs_by_players_t_mid = list()
        
        list_dico_modes_profs_by_players_t_best, \
        list_dico_modes_profs_by_players_t_bad, \
        list_dico_modes_profs_by_players_t_mid\
            = generer_balanced_players_4_modes_profils(
                arr_pl_M_t_vars_modif, 
                m_players, t,
                sum_diff_pos_minus_0_t_minus_2,
                sum_diff_pos_plus_0_t_minus_2,
                sum_cons_is_0_t_minus_2,                             
                sum_prod_is_0_t_minus_2,
                pi_hp_plus, pi_hp_minus,
                a, b,
                pi_0_plus_t, pi_0_minus_t,
                manual_debug, dbg)
          
        list_dico_modes_profs_by_players_t = dict()
        if algo_name == fct_aux.ALGO_NAMES_BF[0]:                              # BEST-BRUTE-FORCE
            list_dico_modes_profs_by_players_t \
                = list_dico_modes_profs_by_players_t_best
            
        elif algo_name == fct_aux.ALGO_NAMES_BF[1]:                            # BAD-BRUTE-FORCE
            list_dico_modes_profs_by_players_t \
                = list_dico_modes_profs_by_players_t_bad
            
        elif algo_name == fct_aux.ALGO_NAMES_BF[2]:                            # MIDDLE-BRUTE-FORCE
            list_dico_modes_profs_by_players_t \
                = list_dico_modes_profs_by_players_t_mid
                
        rd_key = None
        if len(list_dico_modes_profs_by_players_t) == 1:
            rd_key = 0
        else:
            rd_key = np.random.randint(
                        0, 
                        len(list_dico_modes_profs_by_players_t))
        
        id_cpt_xxx, dico_mode_prof_by_players \
            = list_dico_modes_profs_by_players_t[rd_key] 
        print("rd_key={}, cpt_xxx={}".format(rd_key, id_cpt_xxx))
        
        bens_t = dico_mode_prof_by_players["bens_t"]
        csts_t = dico_mode_prof_by_players["csts_t"]
        Perf_t = dico_mode_prof_by_players["Perf_t"]
        b0_t = dico_mode_prof_by_players["b0_t"]
        c0_t = dico_mode_prof_by_players["c0_t"]
        Out_sg = dico_mode_prof_by_players["Out_sg"]
        In_sg = dico_mode_prof_by_players["In_sg"]
        pi_sg_plus_t = dico_mode_prof_by_players["pi_sg_plus_t"]
        pi_sg_minus_t = dico_mode_prof_by_players["pi_sg_minus_t"]
        pi_0_plus_t = dico_mode_prof_by_players["pi_0_plus_t"]
        pi_0_minus_t = dico_mode_prof_by_players["pi_0_minus_t"]
        mode_profile = dico_mode_prof_by_players["mode_profile"]
        
        arr_pl_M_t_vars_modif[:, 
                              fct_aux.AUTOMATE_INDEX_ATTRS["mode_i"]] \
            = mode_profile
        
        print("mode_profile={}, mode_is={}".format(mode_profile, 
                arr_pl_M_t_vars_modif[:,fct_aux.AUTOMATE_INDEX_ATTRS["mode_i"]]))
        print("state_is={} ".format( 
                arr_pl_M_t_vars_modif[:,fct_aux.AUTOMATE_INDEX_ATTRS["state_i"]]))
        
        arr_pl_M_t_vars_modif = balanced_player_game_4_mode_profil(
                                         arr_pl_M_t_vars_modif, 
                                         m_players,
                                         dbg)
        ## test if there are the same values like these in dico_mode_prof_by_players
        In_sg_new, Out_sg_new, b0_t_new, c0_t_new = None, None, None, None
        bens_t_new, csts_t_new = None, None
        pi_sg_plus_t_new, pi_sg_minus_t_new = None, None
        In_sg_new, Out_sg_new, \
        b0_t_new, c0_t_new, \
        bens_t_new, csts_t_new, \
        pi_sg_plus_t_new, pi_sg_minus_t_new \
            = compute_prices_inside_SG(
                    arr_pl_M_t_vars_modif, 
                    sum_diff_pos_minus_0_t_minus_2,
                    sum_diff_pos_plus_0_t_minus_2,
                    sum_cons_is_0_t_minus_2,                             
                    sum_prod_is_0_t_minus_2,
                    pi_hp_plus, pi_hp_minus,
                    a, b,
                    pi_0_plus_t, pi_0_minus_t,
                    manual_debug, dbg)
        bens_csts_t_new = bens_t_new - csts_t_new
        Perf_t_new = np.sum(bens_csts_t_new, axis=0)
        ##### verification of best key quality 
        diff = np.abs(Perf_t_new - Perf_t)
        print(" Perf_t_algo == Perf_t_new --> OK (diff={}) ".format(diff)) \
            if diff < 0.1 \
            else print("Perf_t_algo != Perf_t_new --> NOK (diff={}) \n"\
                       .format(diff))     
        print("b0_t={}, c0_t={}, Out_sg={},In_sg={}, pi_hp_minus_t={}, pi_hp_plus_t={}\n".format(
                    b0_t, c0_t, Out_sg, In_sg, pi_hp_minus_t, pi_hp_plus_t))
        
        # pi_sg_{plus,minus} of shape (T_PERIODS,)
        if np.isnan(pi_sg_plus_t):
            pi_sg_plus_t = 0
        if np.isnan(pi_sg_minus_t):
            pi_sg_minus_t = 0
                
        # checkout NASH equilibrium
        bens_csts_M_t = bens_t - csts_t
        df_nash_t = None
        df_nash_t = checkout_nash_4_profils_by_periods(
                        arr_pl_M_t_vars_modif.copy(),
                        arr_pl_M_T_vars_init[:,t,:],
                        sum_diff_pos_minus_0_t_minus_2,
                        sum_diff_pos_plus_0_t_minus_2,
                        sum_cons_is_0_t_minus_2,                             
                        sum_prod_is_0_t_minus_2,
                        pi_hp_plus, pi_hp_minus,
                        a, b,
                        pi_0_minus_t, pi_0_plus_t, 
                        bens_csts_M_t,
                        m_players,
                        t,
                        manual_debug,
                        dbg)
        df_nash = pd.merge(df_nash, df_nash_t, on='players', how='outer')
        
        ###  complete dico_profils_bf, dico_best_profils_bf, dico_bad_profils_bf: debut
        """
        dico_profils_bf = {"nb_profils":, "profils":[], "Perfs":[]}
        dico_best_profils_bf = {"nb_best_profils":, "profils":[], 
                                "Perfs":[],"nashs":[], "Perfs_nash":[]}
        dico_bad_profils_bf = {"nb_bad_profils":,"profils":[], 
                                "Perfs":[],"nashs":[], "Perfs_nash":[]}
        """
         
        best_profils, best_Perfs \
            = find_best_bad_mid_profils_Perfs(
                list_dico=list_dico_modes_profs_by_players_t_best)
        bad_profils, bad_Perfs \
            = find_best_bad_mid_profils_Perfs(
                list_dico=list_dico_modes_profs_by_players_t_bad)
            
        dico_best_profils_bf["nb_best_profils"] = len(best_profils)
        dico_best_profils_bf["profils"] = best_profils
        dico_best_profils_bf["Perfs"] = best_Perfs
        dico_best_profils_bf["nash"] = []
        dico_best_profils_bf["Perfs_nash"] = []
        
        dico_bad_profils_bf["nb_bad_profils"] = len(bad_profils)
        dico_bad_profils_bf["profils"] = bad_profils
        dico_bad_profils_bf["Perfs"] = bad_Perfs
        dico_bad_profils_bf["nash"] = []
        dico_bad_profils_bf["Perfs_nash"] = []
        
        ##### compute NASH for each X={best, bad} profil
        dico_best_profils_bf = check_nash_4_Xprofils(
                                X_profils=best_profils,
                                dico_X_profils_bf=dico_best_profils_bf,
                                arr_pl_M_t_vars_modif=arr_pl_M_t_vars_modif.copy(),
                                arr_pl_M_T_vars_init=arr_pl_M_T_vars_init,
                                sum_diff_pos_minus_0_t_minus_2=sum_diff_pos_minus_0_t_minus_2,
                                sum_diff_pos_plus_0_t_minus_2=sum_diff_pos_plus_0_t_minus_2,
                                sum_cons_is_0_t_minus_2=sum_cons_is_0_t_minus_2,                             
                                sum_prod_is_0_t_minus_2=sum_prod_is_0_t_minus_2,
                                pi_hp_plus=pi_hp_plus, pi_hp_minus=pi_hp_minus,
                                a=a, b=b,
                                pi_0_plus_t=pi_0_plus_t, pi_0_minus_t=pi_0_minus_t,
                                manual_debug=manual_debug,
                                m_players=m_players,
                                t_periods=t_periods,
                                t=t,
                                dbg=dbg
                                )
        dico_bad_profils_bf = check_nash_4_Xprofils(
                                X_profils=bad_profils,
                                dico_X_profils_bf=dico_bad_profils_bf,
                                arr_pl_M_t_vars_modif=arr_pl_M_t_vars_modif.copy(),
                                arr_pl_M_T_vars_init=arr_pl_M_T_vars_init,
                                sum_diff_pos_minus_0_t_minus_2=sum_diff_pos_minus_0_t_minus_2,
                                sum_diff_pos_plus_0_t_minus_2=sum_diff_pos_plus_0_t_minus_2,
                                sum_cons_is_0_t_minus_2=sum_cons_is_0_t_minus_2,                             
                                sum_prod_is_0_t_minus_2=sum_prod_is_0_t_minus_2,
                                pi_hp_plus=pi_hp_plus, pi_hp_minus=pi_hp_minus,
                                a=a, b=b,
                                pi_0_plus_t=pi_0_plus_t, pi_0_minus_t=pi_0_minus_t,
                                manual_debug=manual_debug,
                                m_players=m_players,
                                t_periods=t_periods,
                                t=t,
                                dbg=dbg
                                )
        ##### compute NASH for each X={best, bad} profil
        
        ###  complete dico_profils_bf, dico_best_profils_bf, dico_bad_profils_bf: fin
        
        #_______     save arr_M_t_vars at t in dataframe : debut    _______
        # df_arr_M_t_vars_modif \
        #     = pd.DataFrame(arr_pl_M_t_vars_modif, 
        #                     columns=fct_aux.AUTOMATE_INDEX_ATTRS.keys(),
        #                     index=dico_id_players["players"])
        # path_to_save_M_t_vars_modif = path_to_save
        # msg = "pi_hp_plus_"+str(pi_hp_plus)+"_pi_hp_minus_"+str(pi_hp_minus)
        # if "simu_DDMM_HHMM" in path_to_save:
        #     path_to_save_M_t_vars_modif \
        #         = os.path.join(name_dir, "simu_"+date_hhmm,
        #                        msg, algo_name, "intermediate_t"
        #                        )
        # Path(path_to_save_M_t_vars_modif).mkdir(parents=True, 
        #                                         exist_ok=True)
            
        # path_2_xls_df_arr_M_t_vars_modif \
        #     = os.path.join(
        #         path_to_save_M_t_vars_modif,
        #           "arr_M_T_vars_{}.xlsx".format(algo_name)
        #         )
        # if not os.path.isfile(path_2_xls_df_arr_M_t_vars_modif):
        #     df_arr_M_t_vars_modif.to_excel(
        #         path_2_xls_df_arr_M_t_vars_modif,
        #         sheet_name="t{}".format(t),
        #         index=True)
        # else:
        #     book = load_workbook(filename=path_2_xls_df_arr_M_t_vars_modif)
        #     with pd.ExcelWriter(path_2_xls_df_arr_M_t_vars_modif, 
        #                         engine='openpyxl') as writer:
        #         writer.book = book
        #         writer.sheets = dict((ws.title, ws) for ws in book.worksheets)    
            
        #         ## Your dataframe to append. 
        #         df_arr_M_t_vars_modif.to_excel(writer, "t{}".format(t))  
            
        #         writer.save() 
        #_______     save arr_M_t_vars at t in dataframe : fin     _______
        
        #___________ update saving variables : debut ______________________
        arr_pl_M_T_vars_modif[:,t,:] = arr_pl_M_t_vars_modif
        b0_ts_T, c0_ts_T, \
        BENs_M_T, CSTs_M_T, \
        pi_sg_plus_T, pi_sg_minus_T, \
        pi_0_plus_T, pi_0_minus_T, \
        df_nash \
            = update_saving_variables(t, 
                b0_ts_T, b0_t,
                c0_ts_T, c0_t,
                BENs_M_T, bens_t,
                CSTs_M_T, csts_t,
                pi_sg_plus_T, pi_sg_plus_t,
                pi_sg_minus_T, pi_sg_minus_t,
                pi_0_plus_T, pi_0_plus_t,
                pi_0_minus_T, pi_0_minus_t,
                df_nash, df_nash
                )
        dico_modes_profs_by_players_t[t] = dico_mode_prof_by_players
        #___________ update saving variables : fin   ______________________
        print("Sis = {}".format(arr_pl_M_T_vars_modif[:,t,fct_aux.AUTOMATE_INDEX_ATTRS["Si"]]))

        print("----- t={} After running free memory={}% ------ ".format(
            t, list(psutil.virtual_memory())[2]))
    
    # __________        compute prices variables         ____________________
    B_is_M, C_is_M, BB_is_M, CC_is_M, EB_is_M, \
    B_is_MT_cum, C_is_MT_cum, \
    B_is_M_T, C_is_M_T, BB_is_M_T, CC_is_M_T, EB_is_M_T \
        = fct_aux.compute_prices_B_C_BB_CC_EB_DET(
                arr_pl_M_T_vars_modif=arr_pl_M_T_vars_modif, 
                pi_sg_minus_T=pi_sg_minus_T, pi_sg_plus_T=pi_sg_plus_T, 
                pi_0_minus_T=pi_0_minus_T, pi_0_plus_T=pi_0_plus_T,
                b0_s_T=b0_ts_T, c0_s_T=c0_ts_T)
        
    dico_EB_R_EBsetA1B1_EBsetB2C = {"EB_setA1B1":[np.nan],"EB_setB2C":[np.nan], 
                                    "ER":[np.nan], "VR":[np.nan]}
    # __________        compute prices variables         ____________________
    
    #_______      save computed variables locally from algo_name     __________
    msg = "pi_hp_plus_"+str(pi_hp_plus)+"_pi_hp_minus_"+str(pi_hp_minus)
    
    print("path_to_save={}".format(path_to_save))
    if "simu_DDMM_HHMM" in path_to_save:
        path_to_save = os.path.join(name_dir, "simu_"+date_hhmm,
                                    msg, algo_name
                                    )
    Path(path_to_save).mkdir(parents=True, exist_ok=True)
    # df_nash.to_excel(os.path.join(
    #             *[path_to_save,
    #               "resume_verify_Nash_equilibrium_{}.xlsx".format(algo_name)]), 
    #             index=False)
    df_nash.to_csv(os.path.join(
                *[path_to_save,
                  "resume_verify_Nash_equilibrium_{}.csv".format(algo_name)]), 
                index=False)

    fct_aux.save_variables(
            path_to_save=path_to_save, 
            arr_pl_M_T_K_vars=arr_pl_M_T_vars_modif, 
            b0_s_T_K=b0_ts_T, c0_s_T_K=c0_ts_T, 
            B_is_M=B_is_M, C_is_M=C_is_M, B_is_M_T=B_is_M_T, C_is_M_T=C_is_M_T,
            BENs_M_T_K=BENs_M_T, CSTs_M_T_K=CSTs_M_T, 
            BB_is_M=BB_is_M, CC_is_M=CC_is_M, EB_is_M=EB_is_M, 
            BB_is_M_T=BB_is_M_T, CC_is_M_T=CC_is_M_T, EB_is_M_T=EB_is_M_T,
            dico_EB_R_EBsetA1B1_EBsetB2C=dico_EB_R_EBsetA1B1_EBsetB2C,
            pi_sg_minus_T_K=pi_sg_minus_T, pi_sg_plus_T_K=pi_sg_plus_T, 
            pi_0_minus_T_K=pi_0_minus_T, pi_0_plus_T_K=pi_0_plus_T,
            pi_hp_plus_T=pi_hp_plus_T, pi_hp_minus_T=pi_hp_minus_T, 
            dico_stats_res=dico_modes_profs_by_players_t, 
            algo=algo_name, 
            dico_best_steps=dict())
    turn_dico_stats_res_into_df_BF(
          dico_modes_profs_players_algo = dico_modes_profs_by_players_t, 
          path_to_save = path_to_save, 
          t_periods = t_periods, 
          manual_debug = manual_debug, 
          algo_name = algo_name)
    
    # _____     variables of matrices of 50 rows and 7 columns      ________
    
    
    return arr_pl_M_T_vars_modif, dico_profils_bf, \
            dico_best_profils_bf, dico_bad_profils_bf
  
    
# __________       main function of One BF Algo   ---> fin        ____________

def find_best_bad_mid_profils_Perfs(list_dico):
    """
    return X_profils and their associated Perf_t s with X = {'best', 'bad', 'mid' some time}

    Parameters
    ----------
    list_dico : list of dictionnary. Each dictionnary looks like 
        {"bens_t":, "csts_t":, "Perf_t": ,"b0_t":, "c0_t":, "mode_profile":, ...}
        
    Returns
    -------
    X_profils, X_Perfs.

    """
    X_profils, X_Perfs = list(), list()
    
    for tupleX in list_dico:
        dico = tupleX[1]
        #print("mode_profile={}".format(dico["mode_profile"]))
        X_profils.append( dico["mode_profile"] )
        X_Perfs.append( dico["Perf_t"] )
    return X_profils, X_Perfs

def check_nash_4_Xprofils(X_profils, dico_X_profils_bf,
                            arr_pl_M_t_vars_modif,
                            arr_pl_M_T_vars_init,
                            sum_diff_pos_minus_0_t_minus_2,
                            sum_diff_pos_plus_0_t_minus_2,
                            sum_cons_is_0_t_minus_2,                             
                            sum_prod_is_0_t_minus_2,
                            pi_hp_plus, pi_hp_minus,
                            a, b,
                            pi_0_plus_t, pi_0_minus_t,
                            manual_debug,
                            m_players,
                            t_periods,
                            t,
                            dbg):
    """
    verify if one profil in X_profils has a Nash equilibrium
    NB: X = {best, bad}
    """
    for X_profil in X_profils:
        arr_M_t_vars_tmp =  arr_pl_M_t_vars_modif.copy()
        arr_M_t_vars_tmp[:,fct_aux.AUTOMATE_INDEX_ATTRS["mode_i"]] = X_profil
        arr_M_t_vars_tmp = balanced_player_game_4_mode_profil(
                                     arr_M_t_vars_tmp, 
                                     m_players,
                                     dbg)
        In_sg_tmp, Out_sg_tmp, b0_t_tmp, c0_t_tmp = None, None, None, None
        bens_t_tmp, csts_t_tmp = None, None
        pi_sg_plus_t_tmp, pi_sg_minus_t_tmp = None, None
        In_sg_tmp, Out_sg_tmp, \
        b0_t_tmp, c0_t_tmp, \
        bens_t_tmp, csts_t_tmp, \
        pi_sg_plus_t_tmp, pi_sg_minus_t_tmp \
            = compute_prices_inside_SG(
                    arr_M_t_vars_tmp, 
                    sum_diff_pos_minus_0_t_minus_2,
                    sum_diff_pos_plus_0_t_minus_2,
                    sum_cons_is_0_t_minus_2,                             
                    sum_prod_is_0_t_minus_2,
                    pi_hp_plus, pi_hp_minus,
                    a, b,
                    pi_0_plus_t, pi_0_minus_t,
                    manual_debug, dbg)
        bens_csts_M_t_tmp = bens_t_tmp - csts_t_tmp
        Perf_t_tmp = np.sum(bens_csts_M_t_tmp, axis=0)
        df_NH_t_tmp = checkout_nash_4_profils_by_periods(
                            arr_M_t_vars_tmp.copy(),
                            arr_pl_M_T_vars_init[:,t,:],
                            sum_diff_pos_minus_0_t_minus_2,
                            sum_diff_pos_plus_0_t_minus_2,
                            sum_cons_is_0_t_minus_2,                             
                            sum_prod_is_0_t_minus_2,
                            pi_hp_plus, pi_hp_minus,
                            a, b,
                            pi_0_minus_t, pi_0_plus_t, 
                            bens_csts_M_t_tmp,
                            m_players,
                            t,
                            manual_debug,
                            dbg)
        bool_equilibrium_nash = all(df_NH_t_tmp.loc[:, "res_t"+str(t_periods-1)] == "STABLE")
        if bool_equilibrium_nash:
            dico_X_profils_bf["nash"].append(X_profil)
            dico_X_profils_bf["Perfs_nash"].append(Perf_t_tmp)
            
    return dico_X_profils_bf



###############################################################################
#                   definition  des unittests
#
###############################################################################
    
# def test_BRUTE_FORCE_balanced_player_game_Pi_Ci_one_period():
    
#     fct_aux.N_DECIMALS = 8
#     a=1; b=1;
#     pi_hp_plus = 10 #0.2*pow(10,-3) #[5, 15]
#     pi_hp_minus = 20 #0.33 #[15, 5]
    
#     manual_debug=False
#     gamma_version = 3 #1,2,3,4
#     debug = False
#     criteria_bf = "Perf_t"
#     used_instances = False #False#True
    
#     setA_m_players = 15; setB_m_players = 10; setC_m_players = 10
#     setA_m_players, setB_m_players, setC_m_players = 8, 3, 3
#     setA_m_players, setB_m_players, setC_m_players = 6, 2, 2
#     t_periods = 1 
#     path_to_arr_pl_M_T = os.path.join(*["tests", "AUTOMATE_INSTANCES_GAMES"])
    
    
#     prob_A_A = 0.7; prob_A_B = 0.3; prob_A_C = 0.0;
#     prob_B_A = 0.3; prob_B_B = 0.4; prob_B_C = 0.3;
#     prob_C_A = 0.1; prob_C_B = 0.2; prob_C_C = 0.7;
#     scenario = [(prob_A_A, prob_A_B, prob_A_C), (prob_B_A, prob_B_B, prob_B_C),
#                 (prob_C_A, prob_C_B, prob_C_C)]
    
#     arr_pl_M_T_vars_init = fct_aux.get_or_create_instance_Pi_Ci_one_period(
#                         setA_m_players, setB_m_players, setC_m_players, 
#                         t_periods, 
#                         scenario,
#                         path_to_arr_pl_M_T, used_instances)
#     fct_aux.checkout_values_Pi_Ci_arr_pl_one_period(arr_pl_M_T_vars_init)
    
#     arr_pl_M_T_vars = bf_balanced_player_game(
#                                 arr_pl_M_T_vars_init.copy(),
#                                 pi_hp_plus=pi_hp_plus, 
#                                 pi_hp_minus=pi_hp_minus,
#                                 a=a, b=b,
#                                 gamma_version = gamma_version,
#                                 path_to_save="tests", 
#                                 name_dir="tests", 
#                                 date_hhmm="DDMM_HHMM",
#                                 manual_debug=manual_debug, 
#                                 criteria_bf=criteria_bf, dbg=debug)
    
#     return arr_pl_M_T_vars
    
def test_BRUTE_FORCE_balanced_player_game_Pi_Ci_t_ONE_ALGO():
    
    fct_aux.N_DECIMALS = 8
    a, b = 1, 1
    pi_hp_plus = 10 #0.2*pow(10,-3) #[5, 15]
    pi_hp_minus = 20 #0.33 #[15, 5]
    
    manual_debug = False
    gamma_version = 2 #1,2
    debug = False
    criteria_bf = "Perf_t"
    used_instances = True #False#True
    
    
    t_periods = 1
    path_to_arr_pl_M_T = os.path.join(*["tests", "AUTOMATE_INSTANCES_GAMES"])
    # _____                     scenarios --> debut                 __________
    scenarios_name = ["scenario1", "scenario2", "scenario3"]  
    scenario_name = scenarios_name[np.random.randint(low=0, high=len(scenarios_name))] 
    prob_scen, scenario, m_players, arr_pl_M_T_vars_init = None, None, None, None
    if scenario_name == "scenario1":
        prob_scen = 0.6
        prob_A_A = prob_scen; prob_A_C = 1-prob_scen;
        prob_C_A = 1-prob_scen; prob_C_C = prob_scen;
        scenario = [(prob_A_A, prob_A_C), 
                    (prob_C_A, prob_C_C)]
        setA_m_players_1 = 10; setC_m_players_1 = 10;                           # 20 joueurs
        setA_m_players_1 = 5; setC_m_players_1 = 5;                             # 10 joueurs
        m_players = setA_m_players_1 + setC_m_players_1
        arr_pl_M_T_vars_init \
            = fct_aux.get_or_create_instance_Pi_Ci_etat_AUTOMATE_SETAC_doc23(
                            setA_m_players_1, setC_m_players_1, 
                            t_periods, 
                            scenario,
                            scenario_name,
                            path_to_arr_pl_M_T, used_instances)
        fct_aux.checkout_values_Pi_Ci_arr_pl_SETAC_doc23(arr_pl_M_T_vars_init, 
                                                         scenario_name)
        
    elif scenario_name in ["scenario2", "scenario3"]:
        prob_scen = 0.8 if scenario_name == "scenario2" else 0.5
        prob_A_A = 1-prob_scen; prob_A_B1 = prob_scen; prob_A_B2 = 0.0; prob_A_C = 0.0;
        prob_B1_A = prob_scen; prob_B1_B1 = 1-prob_scen; prob_B1_B2 = 0.0; prob_B1_C = 0.0;
        prob_B2_A = 0.0; prob_B2_B1 = 0.0; prob_B2_B2 = 1-prob_scen; prob_B2_C = prob_scen;
        prob_C_A = 0.0; prob_C_B1 = 0.0; prob_C_B2 = prob_scen; prob_C_C = 1-prob_scen 
        scenario = [(prob_A_A, prob_A_B1, prob_A_B2, prob_A_C), 
                    (prob_B1_A, prob_B1_B1, prob_B1_B2, prob_B1_C),
                    (prob_B2_A, prob_B2_B1, prob_B2_B2, prob_B2_C),
                    (prob_C_A, prob_C_B1, prob_C_B2, prob_C_C)]
        setA_m_players_23 = 10; setB1_m_players_23 = 3; 
        setB2_m_players_23 = 5; setC_m_players_23 = 8; 
        setA_m_players_23 = 5; setB1_m_players_23 = 2; 
        setB2_m_players_23 = 2; setC_m_players_23 = 4;                          # 13 joueurs
        m_players = setA_m_players_23 + setB1_m_players_23 \
                    + setB2_m_players_23 + setC_m_players_23
        arr_pl_M_T_vars_init \
            = fct_aux.get_or_create_instance_Pi_Ci_etat_AUTOMATE_SETAB1B2C_doc23(
                        setA_m_players_23, setB1_m_players_23, 
                        setB2_m_players_23, setC_m_players_23, 
                        t_periods, 
                        scenario,
                        scenario_name,
                        path_to_arr_pl_M_T, used_instances)
        fct_aux.checkout_values_Pi_Ci_arr_pl_SETAB1B2C_doc23(arr_pl_M_T_vars_init, 
                                                             scenario_name) 
     
    # _____                     scenarios --> fin                   __________
    
    global MOD
    MOD = int(0.10*pow(2, m_players)) \
            if pow(2, m_players) < 65000 \
            else int(0.020*pow(2, m_players))    
    
    algo_name = fct_aux.ALGO_NAMES_BF[0]
    name_simu = algo_name+"_simu_"+datetime.now().strftime("%d%m_%H%M")
    path_to_save = os.path.join("tests", name_simu)
    
    arr_pl_M_T_vars, dico_profils_bf, \
    dico_best_profils_bf, dico_bad_profils_bf \
        = bf_balanced_player_game_ONE_ALGO(
                        arr_pl_M_T_vars_init.copy(),
                        algo_name,
                        pi_hp_plus=pi_hp_plus, 
                        pi_hp_minus=pi_hp_minus,
                        a=a, b=b,
                        gamma_version = gamma_version,
                        path_to_save=path_to_save, 
                        name_dir="tests", 
                        date_hhmm="DDMM_HHMM",
                        manual_debug=manual_debug, 
                        criteria_bf=criteria_bf, dbg=debug)
    
    return arr_pl_M_T_vars, dico_profils_bf, \
            dico_best_profils_bf, dico_bad_profils_bf

def test_BRUTE_FORCE_balanced_player_game_Pi_Ci_one_period_ONE_ALGO():
    
    fct_aux.N_DECIMALS = 8
    a, b = 1, 1
    pi_hp_plus = 10 #0.2*pow(10,-3) #[5, 15]
    pi_hp_minus = 20 #0.33 #[15, 5]
    
    manual_debug = False
    gamma_version = -2  #-2,-1,1,2,3,4,5
    debug = False
    criteria_bf = "Perf_t"
    used_instances = True #False#True
    
    t_periods = 1
    path_to_arr_pl_M_T = os.path.join(*["tests", "AUTOMATE_INSTANCES_GAMES"])
    
    # _____                     scenarios --> debut                 __________
    scenario_name = "scenarioOnePeriod"      
    setA_m_players, setB_m_players, setC_m_players = 6, 3, 3                   # 12 players
    arr_pl_M_T_vars_init \
        = fct_aux.get_or_create_instance_Pi_Ci_one_period_doc23(
            setA_m_players=setA_m_players, 
            setB_m_players=setB_m_players, 
            setC_m_players=setC_m_players, 
            t_periods=t_periods, 
            scenario=None,
            scenario_name=scenario_name,
            path_to_arr_pl_M_T=path_to_arr_pl_M_T, 
            used_instances=used_instances)
    # _____                     scenarios --> fin                   __________
    
    global MOD
    m_players = setA_m_players + setB_m_players + setC_m_players
    MOD = int(0.10*pow(2, m_players)) \
            if pow(2, m_players) < 65000 \
            else int(0.020*pow(2, m_players))    
    
    algo_name = fct_aux.ALGO_NAMES_BF[0]
    name_simu = algo_name+"_simu_"+datetime.now().strftime("%d%m_%H%M")
    path_to_save = os.path.join("tests", name_simu)
    
    arr_pl_M_T_vars, dico_profils_bf, \
    dico_best_profils_bf, dico_bad_profils_bf \
        = bf_balanced_player_game_ONE_ALGO(
                        arr_pl_M_T_vars_init.copy(),
                        algo_name,
                        pi_hp_plus=pi_hp_plus, 
                        pi_hp_minus=pi_hp_minus,
                        a=a, b=b,
                        gamma_version = gamma_version,
                        path_to_save=path_to_save, 
                        name_dir="tests", 
                        date_hhmm="DDMM_HHMM",
                        manual_debug=manual_debug, 
                        criteria_bf=criteria_bf, dbg=debug)
    
    return arr_pl_M_T_vars, dico_profils_bf, \
            dico_best_profils_bf, dico_bad_profils_bf

###############################################################################
#                   Execution
#
###############################################################################
if __name__ == "__main__":
    ti = time.time()
    
    # Celui la est incorrect parce que je ne sais pas utiliser le stock a t-1 sur une algo combinant les 3 algorithmes
    # arr_pl_M_T_vars_modif \
    #     = test_BRUTE_FORCE_balanced_player_game_Pi_Ci_NEW_AUTOMATE()
       
    if np.random.randint(low=0, high=2):
        arr_pl_M_T_vars_modif, dico_profils_bf, \
        dico_best_profils_bf, dico_bad_profils_bf \
            = test_BRUTE_FORCE_balanced_player_game_Pi_Ci_t_ONE_ALGO()
    else:
        arr_pl_M_T_vars, dico_profils_bf, \
        dico_best_profils_bf, dico_bad_profils_bf \
            = test_BRUTE_FORCE_balanced_player_game_Pi_Ci_one_period_ONE_ALGO()
    
    print("runtime = {}".format(time.time() - ti))